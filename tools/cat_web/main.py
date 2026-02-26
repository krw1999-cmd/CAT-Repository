from __future__ import annotations

"""CAT Web — Flask application entry point and route definitions."""

import csv
import io
import json
import os
import uuid
from datetime import date, datetime, timedelta

from flask import (
    Flask,
    Response,
    g,
    make_response,
    redirect,
    render_template,
    request,
    session,
    url_for,
    flash,
    abort,
    jsonify,
)
from flask_login import login_required, login_user, logout_user, current_user

import auth
import config
import db as db_module
import drive_sync

app = Flask(__name__)
app.secret_key = config.SECRET_KEY
app.config["DEBUG"] = config.DEBUG

auth.init_app(app)
app.teardown_appcontext(db_module.close_db)

TRANSACTION_TYPES = db_module.TRANSACTION_TYPES
ASSIGNEE_ROLES = db_module.ASSIGNEE_ROLES


@app.template_filter("money")
def money_filter(value):
    try:
        return "${:,.2f}".format(float(value or 0))
    except (TypeError, ValueError):
        return "$0.00"


def _clean_date(s) -> str:
    """Return s if it is a valid YYYY-MM-DD date, otherwise empty string."""
    if not s:
        return ""
    s = str(s).strip()
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return s
    except ValueError:
        return ""


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("claims_list"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user_row = db_module.get_user_by_username(username)
        if user_row and auth.check_password(user_row["password_hash"], password):
            user = auth.User(user_row["id"], user_row["username"])
            login_user(user, remember=True)
            return redirect(request.args.get("next") or url_for("claims_list"))
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# Claims dashboard
# ---------------------------------------------------------------------------

@app.route("/")
@login_required
def claims_list():
    claims = db_module.get_all_claims()
    return render_template("claims.html", claims=claims)


@app.route("/claims/new", methods=["GET", "POST"])
@login_required
def new_claim():
    if request.method == "POST":
        data = {
            "job_number": request.form.get("job_number", "").strip(),
            "claim_number": request.form.get("claim_number", "").strip(),
            "insured_name": request.form.get("insured_name", "").strip(),
            "carrier": request.form.get("carrier", "").strip(),
            "contract_pct": db_module._f(request.form.get("contract_pct")),
        }
        claim_id = db_module.create_claim(data)
        return redirect(url_for("claim_detail", claim_id=claim_id))
    return render_template("new_claim.html")


# ---------------------------------------------------------------------------
# Claim detail
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>")
@login_required
def claim_detail(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim:
        abort(404)
    limits = db_module.get_limits(claim_id)
    transactions = db_module.get_transactions(claim_id)
    expenses = db_module.get_expenses(claim_id)
    assignees = db_module.get_assignees(claim_id)
    summary = db_module.get_transaction_summary(claim_id)
    open_escrows = db_module.get_open_escrows(claim_id)
    coverage_pivot = db_module.get_coverage_pivot(claim_id)
    vendor_activity = db_module.get_claim_vendor_activity(claim_id)
    return render_template(
        "claim.html",
        claim=claim,
        limits=limits,
        transactions=transactions,
        expenses=expenses,
        assignees=assignees,
        summary=summary,
        open_escrows=open_escrows,
        coverage_pivot=coverage_pivot,
        vendor_activity=vendor_activity,
        tx_types=TRANSACTION_TYPES,
        coverage_types=db_module.get_claim_coverage_types(claim_id),
    )


# -- Claim header edit -------------------------------------------------------

@app.route("/claims/<int:claim_id>/header")
@login_required
def claim_header_view(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim:
        abort(404)
    return render_template("partials/_claim_header.html", claim=claim)


@app.route("/claims/<int:claim_id>/header/edit")
@login_required
def claim_header_edit(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim:
        abort(404)
    return render_template("partials/_claim_header_edit.html", claim=claim)


@app.route("/claims/<int:claim_id>/header", methods=["POST"])
@login_required
def claim_header_save(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim:
        abort(404)
    data = {
        "job_number": request.form.get("job_number", "").strip(),
        "claim_number": request.form.get("claim_number", "").strip(),
        "insured_name": request.form.get("insured_name", "").strip(),
        "carrier": request.form.get("carrier", "").strip(),
        "contract_pct": db_module._f(request.form.get("contract_pct")),
        "mortgage_company": request.form.get("mortgage_company", "").strip() or None,
        "notes": request.form.get("notes", "").strip() or None,
        "proc_method_carrier": request.form.get("proc_method_carrier") or None,
        "proc_method_draw":    request.form.get("proc_method_draw") or None,
        "proc_method_escrow":  request.form.get("proc_method_escrow") or None,
        "drive_folder_url":    request.form.get("drive_folder_url", "").strip() or None,
    }
    db_module.update_claim(claim_id, data)
    claim = db_module.get_claim(claim_id)
    return render_template("partials/_claim_header.html", claim=claim)


@app.route("/claims/<int:claim_id>/sync-drive", methods=["POST"])
@login_required
def claim_sync_drive(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim or not claim.get("drive_folder_url"):
        return ("No folder URL set", 400)

    files = drive_sync.list_folder_files(claim["drive_folder_url"])

    txs  = db_module.get_transactions(claim_id)
    exps = db_module.get_expenses(claim_id)

    tx_links  = {}
    exp_links = {}
    matched = 0

    for tx in txs:
        if (tx.get("type") or "").lower() == "fee invoice":
            url = drive_sync.match_invoice(files, tx.get("invoice_number") or "")
        else:
            url = drive_sync.match_check(files, tx.get("check_number") or "")
        if url:
            tx_links[tx["id"]] = url
            matched += 1

    for exp in exps:
        url = drive_sync.match_invoice(files, exp.get("invoice_number") or "")
        if url:
            exp_links[exp["id"]] = url
            matched += 1

    db_module.sync_drive_links(claim_id, tx_links, exp_links)
    return f"Synced {matched} file(s)", 200


# ---------------------------------------------------------------------------
# Policy Limits — HTMX fragments
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/limits/new-row")
@login_required
def limit_new_row(claim_id: int):
    claim = db_module.get_claim(claim_id)
    return render_template("partials/_limit_edit.html",
                           claim=claim,
                           limit=None,
                           coverage_types=db_module.get_claim_coverage_types(claim_id))


@app.route("/claims/<int:claim_id>/limits", methods=["POST"])
@login_required
def limit_create(claim_id: int):
    data = _limit_data_from_form()
    db_module.ensure_coverage_type(data["coverage_type"])
    limit_id = db_module.create_limit(claim_id, data)
    limit = db_module.get_limit(limit_id)
    claim = db_module.get_claim(claim_id)
    row_html = render_template("partials/_limit_row.html", claim=claim, limit=limit)
    placeholder = '<tr id="limit-new-row-placeholder"></tr>'
    return row_html + placeholder


@app.route("/claims/<int:claim_id>/limits/<int:limit_id>/edit")
@login_required
def limit_edit_row(claim_id: int, limit_id: int):
    claim = db_module.get_claim(claim_id)
    limit = db_module.get_limit(limit_id)
    if not limit:
        abort(404)
    return render_template("partials/_limit_edit.html",
                           claim=claim,
                           limit=limit,
                           coverage_types=db_module.get_claim_coverage_types(claim_id))


@app.route("/claims/<int:claim_id>/limits/<int:limit_id>", methods=["PUT"])
@login_required
def limit_update(claim_id: int, limit_id: int):
    data = _limit_data_from_form()
    db_module.ensure_coverage_type(data["coverage_type"])
    db_module.update_limit(limit_id, data)
    limit = db_module.get_limit(limit_id)
    claim = db_module.get_claim(claim_id)
    return render_template("partials/_limit_row.html", claim=claim, limit=limit)


@app.route("/claims/<int:claim_id>/limits/<int:limit_id>", methods=["DELETE"])
@login_required
def limit_delete(claim_id: int, limit_id: int):
    db_module.delete_limit(limit_id)
    return ""


def _limit_data_from_form() -> dict:
    return {
        "coverage_type": request.form.get("coverage_type", "").strip(),
        "base_limit": db_module._f(request.form.get("base_limit")),
        "extended_limit": db_module._f(request.form.get("extended_limit")),
    }


# ---------------------------------------------------------------------------
# Transactions — HTMX fragments
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/new-row")
@login_required
def tx_new_row(claim_id: int):
    claim = db_module.get_claim(claim_id)
    return render_template("partials/_tx_edit.html",
                           claim=claim,
                           tx=None,
                           tx_types=TRANSACTION_TYPES)


@app.route("/claims/<int:claim_id>/transactions", methods=["POST"])
@login_required
def tx_create(claim_id: int):
    data = _tx_data_from_form()
    tx_id = db_module.create_transaction(claim_id, data)
    tx = db_module.get_transactions(claim_id)  # use joined version for display
    tx_row = next((t for t in tx if t["id"] == tx_id), None)
    claim = db_module.get_claim(claim_id)
    summary = db_module.get_transaction_summary(claim_id)
    if tx_row is None:
        tx_row = db_module.get_transaction(tx_id)
    row_html = render_template("partials/_tx_row.html", claim=claim, tx=tx_row,
                               tx_types=TRANSACTION_TYPES)
    placeholder = '<tr id="tx-new-row-placeholder"></tr>'
    summary_html = render_template("partials/_summary.html", summary=summary)
    return row_html + placeholder + summary_html


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/edit")
@login_required
def tx_edit_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    if not tx:
        abort(404)
    return render_template("partials/_tx_edit.html",
                           claim=claim,
                           tx=tx,
                           tx_types=TRANSACTION_TYPES)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>", methods=["PUT"])
@login_required
def tx_update(claim_id: int, tx_id: int):
    tx_existing = db_module.get_transaction(tx_id)
    if not tx_existing:
        abort(404)
    # Patch: start from current DB values so detail-page fields aren't wiped
    data = dict(tx_existing)
    data.update({
        "check_number":     request.form.get("check_number", "").strip() or None,
        "date":             _clean_date(request.form.get("date")),
        "type":             request.form.get("type", "").strip(),
        "amount":           db_module._f(request.form.get("amount")),
        "payer":            request.form.get("payer", "").strip(),
        "notes":            request.form.get("notes", "").strip(),
        "linked_escrow_id": request.form.get("linked_escrow_id") or None,
    })
    db_module.update_transaction(tx_id, data)
    txs = db_module.get_transactions(claim_id)
    tx = next((t for t in txs if t["id"] == tx_id), None)
    if tx is None:
        tx = db_module.get_transaction(tx_id)
    claim = db_module.get_claim(claim_id)
    summary = db_module.get_transaction_summary(claim_id)
    row_html = render_template("partials/_tx_row.html", claim=claim, tx=tx,
                               tx_types=TRANSACTION_TYPES)
    summary_html = render_template("partials/_summary.html", summary=summary)
    return row_html + summary_html


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>", methods=["DELETE"])
@login_required
def tx_delete(claim_id: int, tx_id: int):
    claim_id_check = db_module.get_transaction(tx_id)
    if not claim_id_check:
        abort(404)
    db_module.delete_transaction(tx_id)
    summary = db_module.get_transaction_summary(claim_id)
    return render_template("partials/_summary.html", summary=summary)


def _tx_data_from_form() -> dict:
    return {
        "sequence_number":     request.form.get("sequence_number") or None,
        "check_number":        request.form.get("check_number", "").strip() or None,
        "date":                _clean_date(request.form.get("date")),
        "amount":              db_module._f(request.form.get("amount")),
        "type":                request.form.get("type", "").strip(),
        "fee_owed":            db_module._f(request.form.get("fee_owed")),
        "balance":             db_module._f(request.form.get("balance")),
        "deferred":            db_module._f(request.form.get("deferred")),
        "recouped":            db_module._f(request.form.get("recouped")),
        "fee_collected":       db_module._f(request.form.get("fee_collected")),
        "reimbursed":          db_module._f(request.form.get("reimbursed")),
        "total_collected":     db_module._f(request.form.get("total_collected")),
        "unpaid_payee_expense":db_module._f(request.form.get("unpaid_payee_expense")),
        "outstanding_expense": db_module._f(request.form.get("outstanding_expense")),
        "ott":                 db_module._fopt(request.form.get("ott")),
        "notes":               request.form.get("notes", "").strip(),
        "check_id":            request.form.get("check_id") or None,
        "linked_escrow_id":    request.form.get("linked_escrow_id") or None,
    }


# ---------------------------------------------------------------------------
# Expenses — HTMX fragments
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/expenses/new-row")
@login_required
def exp_new_row(claim_id: int):
    claim = db_module.get_claim(claim_id)
    vendors = db_module.get_all_vendors()
    return render_template("partials/_exp_edit.html", claim=claim, exp=None, vendors=vendors)


@app.route("/claims/<int:claim_id>/expenses", methods=["POST"])
@login_required
def exp_create(claim_id: int):
    data = _exp_data_from_form()
    exp_id = db_module.create_expense(claim_id, data)
    exp = db_module.get_expense(exp_id)
    claim = db_module.get_claim(claim_id)
    vendors = db_module.get_all_vendors()
    row_html = render_template("partials/_exp_row.html", claim=claim, exp=exp)
    placeholder = '<tr id="exp-new-row-placeholder"></tr>'
    return row_html + placeholder


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/edit")
@login_required
def exp_edit_row(claim_id: int, exp_id: int):
    claim = db_module.get_claim(claim_id)
    exp = db_module.get_expense(exp_id)
    vendors = db_module.get_all_vendors()
    if not exp:
        abort(404)
    return render_template("partials/_exp_edit.html", claim=claim, exp=exp, vendors=vendors)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>", methods=["PUT"])
@login_required
def exp_update(claim_id: int, exp_id: int):
    data = _exp_data_from_form()
    db_module.update_expense(exp_id, data)
    exp = db_module.get_expense(exp_id)
    claim = db_module.get_claim(claim_id)
    return render_template("partials/_exp_row.html", claim=claim, exp=exp)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>", methods=["DELETE"])
@login_required
def exp_delete(claim_id: int, exp_id: int):
    db_module.delete_expense(exp_id)
    return ""


def _exp_data_from_form() -> dict:
    return {
        "invoice_date":      _clean_date(request.form.get("invoice_date")),
        "payee_name":        request.form.get("payee_name", "").strip(),
        "vendor_id":         request.form.get("vendor_id") or None,
        "invoice_amount":    db_module._f(request.form.get("invoice_amount")),
        "responsible_party": request.form.get("responsible_party", "").strip(),
        "unpaid_to_payee":   db_module._f(request.form.get("unpaid_to_payee")),
        "client_outstanding":db_module._f(request.form.get("client_outstanding")),
        "wp_outstanding":    db_module._f(request.form.get("wp_outstanding")),
        "reason":            request.form.get("reason", "").strip(),
        "invoice_number":    request.form.get("invoice_number", "").strip(),
        "client_amount":     db_module._f(request.form.get("client_amount")),
        "wp_amount":         db_module._f(request.form.get("wp_amount")),
    }


# ---------------------------------------------------------------------------
# Expense detail page
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>")
@login_required
def expense_detail(claim_id: int, exp_id: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    if not claim or not expense:
        abort(404)
    payments           = db_module.get_expense_payments(exp_id)
    reimbursements     = db_module.get_expense_reimbursements(exp_id)
    wp_reimbursements  = db_module.get_wp_reimbursements(exp_id)
    splits             = db_module.get_expense_splits(exp_id)
    totals             = db_module.get_expense_totals(exp_id)
    claim_txs          = _vendor_txs_for_expense(claim_id, expense)
    vendors            = db_module.get_all_vendors()
    return render_template(
        "expense.html",
        claim=claim, expense=expense,
        payments=payments, reimbursements=reimbursements,
        wp_reimbursements=wp_reimbursements,
        splits=splits, totals=totals,
        claim_txs_for_select=claim_txs,
        vendors=vendors,
    )


# Sidebar fragment — triggered via HX-Trigger: expTotalsUpdated from mutation routes

@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/sidebar")
@login_required
def exp_sidebar_fragment(claim_id: int, exp_id: int):
    expense = db_module.get_expense(exp_id)
    totals  = db_module.get_expense_totals(exp_id)
    return render_template("partials/_exp_sidebar_content.html",
                           expense=expense, totals=totals)


# Expense info card — view / edit / save

@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/info")
@login_required
def exp_info_view(claim_id: int, exp_id: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    if not expense:
        abort(404)
    return render_template("partials/_exp_info.html", claim=claim, expense=expense)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/info/edit")
@login_required
def exp_info_edit(claim_id: int, exp_id: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    vendors = db_module.get_all_vendors()
    if not expense:
        abort(404)
    return render_template("partials/_exp_info_edit.html", claim=claim, expense=expense, vendors=vendors)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/info", methods=["POST"])
@login_required
def exp_info_save(claim_id: int, exp_id: int):
    data = _exp_data_from_form()
    db_module.update_expense(exp_id, data)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    return render_template("partials/_exp_info.html", claim=claim, expense=expense)


def _vendor_txs_for_expense(claim_id: int, expense: dict) -> list[dict]:
    """Return check transactions filtered to only those with a disbursement to this expense's vendor."""
    vid = expense.get("vendor_id") if expense else None
    if not vid:
        return []
    return db_module.get_claim_txs_with_vendor_disbursement(claim_id, int(vid))


# Expense payments

@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/payments/new-row")
@login_required
def exp_pay_new_row(claim_id: int, exp_id: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    return render_template("partials/_exp_pay_edit.html", claim=claim, expense=expense, p=None)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/payments", methods=["POST"])
@login_required
def exp_pay_create(claim_id: int, exp_id: int):
    data = {
        "date":         (_clean_date(request.form.get("date")) or ""),
        "paid_by":      request.form.get("paid_by", "wp"),
        "amount":       db_module._f(request.form.get("amount")),
        "method":       request.form.get("method", "").strip(),
        "reference":    request.form.get("reference", "").strip(),
        "linked_tx_id": request.form.get("linked_tx_id") or None,
        "notes":        request.form.get("notes", "").strip(),
    }
    pid     = db_module.create_expense_payment(exp_id, data)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    p       = _row_from_list(db_module.get_expense_payments(exp_id), pid)
    claim_txs = _vendor_txs_for_expense(claim_id, expense)
    row_html    = render_template("partials/_exp_pay_row.html", claim=claim, expense=expense, p=p,
                                  claim_txs_for_select=claim_txs)
    placeholder = '<tr id="exp-pay-new-row-placeholder"></tr>'
    resp = make_response(row_html + placeholder)
    resp.headers["HX-Trigger"] = "expTotalsUpdated"
    return resp


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/payments/<int:pid>/edit")
@login_required
def exp_pay_edit_row(claim_id: int, exp_id: int, pid: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    p       = _row_from_list(db_module.get_expense_payments(exp_id), pid)
    return render_template("partials/_exp_pay_edit.html", claim=claim, expense=expense, p=p)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/payments/<int:pid>/view")
@login_required
def exp_pay_view_row(claim_id: int, exp_id: int, pid: int):
    claim     = db_module.get_claim(claim_id)
    expense   = db_module.get_expense(exp_id)
    p         = _row_from_list(db_module.get_expense_payments(exp_id), pid)
    claim_txs = _vendor_txs_for_expense(claim_id, expense)
    return render_template("partials/_exp_pay_row.html", claim=claim, expense=expense, p=p,
                           claim_txs_for_select=claim_txs)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/payments/<int:pid>", methods=["PUT"])
@login_required
def exp_pay_update(claim_id: int, exp_id: int, pid: int):
    data = {
        "date":         (_clean_date(request.form.get("date")) or ""),
        "paid_by":      request.form.get("paid_by", "wp"),
        "amount":       db_module._f(request.form.get("amount")),
        "method":       request.form.get("method", "").strip(),
        "reference":    request.form.get("reference", "").strip(),
        "linked_tx_id": request.form.get("linked_tx_id") or None,
        "notes":        request.form.get("notes", "").strip(),
    }
    db_module.update_expense_payment(pid, data)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    p       = _row_from_list(db_module.get_expense_payments(exp_id), pid)
    claim_txs = _vendor_txs_for_expense(claim_id, expense)
    row_html = render_template("partials/_exp_pay_row.html", claim=claim, expense=expense, p=p,
                               claim_txs_for_select=claim_txs)
    resp = make_response(row_html)
    resp.headers["HX-Trigger"] = "expTotalsUpdated"
    return resp


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/payments/<int:pid>", methods=["DELETE"])
@login_required
def exp_pay_delete(claim_id: int, exp_id: int, pid: int):
    db_module.delete_expense_payment(pid)
    resp = make_response("")
    resp.headers["HX-Trigger"] = "expTotalsUpdated"
    return resp


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/payments/<int:pid>/link-tx", methods=["POST"])
@login_required
def exp_pay_link_tx(claim_id: int, exp_id: int, pid: int):
    raw = request.form.get("tx_id") or ""
    tx_id = int(raw) if raw.strip() else None
    db_module.link_expense_payment_to_tx(pid, tx_id)
    claim     = db_module.get_claim(claim_id)
    expense   = db_module.get_expense(exp_id)
    p         = _row_from_list(db_module.get_expense_payments(exp_id), pid)
    claim_txs = _vendor_txs_for_expense(claim_id, expense)
    return render_template("partials/_exp_pay_row.html", claim=claim, expense=expense, p=p,
                           claim_txs_for_select=claim_txs)


# Expense reimbursements

@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/reimbursements/new-row")
@login_required
def exp_reimb_new_row(claim_id: int, exp_id: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    claim_txs = db_module.get_claim_transactions_for_select(claim_id)
    return render_template("partials/_exp_reimb_edit.html",
                           claim=claim, expense=expense, r=None,
                           claim_txs_for_select=claim_txs)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/reimbursements", methods=["POST"])
@login_required
def exp_reimb_create(claim_id: int, exp_id: int):
    data = {
        "date":        (_clean_date(request.form.get("date")) or ""),
        "amount":      db_module._f(request.form.get("amount")),
        "method":      request.form.get("method", "").strip(),
        "reference":   request.form.get("reference", "").strip(),
        "linked_tx_id":request.form.get("linked_tx_id") or None,
        "notes":       request.form.get("notes", "").strip(),
    }
    rid     = db_module.create_expense_reimbursement(exp_id, data)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    r       = _row_from_list(db_module.get_expense_reimbursements(exp_id), rid)
    row_html    = render_template("partials/_exp_reimb_row.html", claim=claim, expense=expense, r=r)
    placeholder = '<tr id="exp-reimb-new-row-placeholder"></tr>'
    resp = make_response(row_html + placeholder)
    resp.headers["HX-Trigger"] = "expTotalsUpdated"
    return resp


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/reimbursements/<int:rid>", methods=["DELETE"])
@login_required
def exp_reimb_delete(claim_id: int, exp_id: int, rid: int):
    db_module.delete_expense_reimbursement(rid)
    resp = make_response("")
    resp.headers["HX-Trigger"] = "expTotalsUpdated"
    return resp


# Expense WP reimbursements (WP paying client back)

@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/wp-reimbursements/new-row")
@login_required
def exp_wp_reimb_new_row(claim_id: int, exp_id: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    claim_txs = db_module.get_claim_transactions_for_select(claim_id)
    return render_template("partials/_exp_wp_reimb_edit.html",
                           claim=claim, expense=expense, r=None,
                           claim_txs_for_select=claim_txs)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/wp-reimbursements", methods=["POST"])
@login_required
def exp_wp_reimb_create(claim_id: int, exp_id: int):
    data = {
        "date":      (_clean_date(request.form.get("date")) or ""),
        "amount":    db_module._f(request.form.get("amount")),
        "method":    request.form.get("method", "").strip(),
        "reference": request.form.get("reference", "").strip(),
        "notes":     request.form.get("notes", "").strip(),
    }
    rid     = db_module.create_wp_reimbursement(exp_id, data)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    r       = _row_from_list(db_module.get_wp_reimbursements(exp_id), rid)
    row_html    = render_template("partials/_exp_wp_reimb_row.html", claim=claim, expense=expense, r=r)
    placeholder = '<tr id="exp-wp-reimb-new-row-placeholder"></tr>'
    resp = make_response(row_html + placeholder)
    resp.headers["HX-Trigger"] = "expTotalsUpdated"
    return resp


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/wp-reimbursements/<int:rid>", methods=["DELETE"])
@login_required
def exp_wp_reimb_delete(claim_id: int, exp_id: int, rid: int):
    db_module.delete_wp_reimbursement(rid)
    resp = make_response("")
    resp.headers["HX-Trigger"] = "expTotalsUpdated"
    return resp


# Expense splits

@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/splits/new-row")
@login_required
def exp_split_new_row(claim_id: int, exp_id: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    return render_template("partials/_exp_split_edit.html",
                           claim=claim, expense=expense, split=None,
                           roles=db_module.ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/splits/<int:sid>/edit")
@login_required
def exp_split_edit_row(claim_id: int, exp_id: int, sid: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    splits  = db_module.get_expense_splits(exp_id)
    split   = _row_from_list(splits, sid)
    if not split:
        abort(404)
    return render_template("partials/_exp_split_edit.html",
                           claim=claim, expense=expense, split=split,
                           roles=db_module.ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/splits/<int:sid>/view")
@login_required
def exp_split_view_row(claim_id: int, exp_id: int, sid: int):
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    splits  = db_module.get_expense_splits(exp_id)
    split   = _row_from_list(splits, sid)
    if not split:
        abort(404)
    return render_template("partials/_exp_split_row.html", claim=claim, expense=expense, split=split)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/splits", methods=["POST"])
@login_required
def exp_split_create(claim_id: int, exp_id: int):
    rid = request.form.get("recipient_id") or None
    data = {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "recipient_id": rid,
    }
    sid     = db_module.create_expense_split(exp_id, data)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    split   = _row_from_list(db_module.get_expense_splits(exp_id), sid)
    row_html = render_template("partials/_exp_split_row.html", claim=claim, expense=expense, split=split)
    placeholder = '<tr id="exp-split-new-row-placeholder"></tr>'
    return row_html + placeholder


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/splits/<int:sid>", methods=["PUT"])
@login_required
def exp_split_update(claim_id: int, exp_id: int, sid: int):
    rid = request.form.get("recipient_id") or None
    data = {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "recipient_id": rid,
    }
    db_module.update_expense_split(sid, data)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    split   = _row_from_list(db_module.get_expense_splits(exp_id), sid)
    return render_template("partials/_exp_split_row.html", claim=claim, expense=expense, split=split)


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/splits/<int:sid>", methods=["DELETE"])
@login_required
def exp_split_delete(claim_id: int, exp_id: int, sid: int):
    db_module.delete_expense_split(sid)
    return ""


@app.route("/claims/<int:claim_id>/expenses/<int:exp_id>/splits/seed", methods=["POST"])
@login_required
def exp_split_seed(claim_id: int, exp_id: int):
    # Delete existing and re-clone
    for s in db_module.get_expense_splits(exp_id):
        db_module.delete_expense_split(s["id"])
    db_module.clone_splits_from_assignees_for_expense(claim_id, exp_id)
    claim   = db_module.get_claim(claim_id)
    expense = db_module.get_expense(exp_id)
    splits  = db_module.get_expense_splits(exp_id)
    return render_template("partials/_exp_splits_section.html", claim=claim, expense=expense, splits=splits)


def _row_from_list(rows: list, row_id: int) -> dict | None:
    for r in rows:
        if r.get("id") == row_id:
            return r
    return None


# ---------------------------------------------------------------------------
# Check queue (Phase 3 bridge)
# ---------------------------------------------------------------------------

@app.route("/checks")
@login_required
def checks_queue():
    checks = db_module.get_unmatched_checks()
    all_claims = db_module.get_all_claims_simple()
    return render_template("checks.html", checks=checks, all_claims=all_claims)


@app.route("/checks/<int:check_id>/match", methods=["POST"])
@login_required
def check_match(check_id: int):
    claim_id = request.form.get("claim_id")
    if not claim_id:
        flash("Select a claim first.", "error")
        return redirect(url_for("checks_queue"))
    claim_id = int(claim_id)
    db_module.match_check_to_claim(check_id, claim_id)

    from db import get_db
    check_row = get_db().execute("SELECT * FROM checks WHERE id = ?", (check_id,)).fetchone()
    if check_row:
        check_row = dict(check_row)
        claim = db_module.get_claim(claim_id)
        pct = (claim["contract_pct"] or 0) / 100.0
        try:
            amount = float(str(check_row.get("amount", "0")).lstrip("$").replace(",", ""))
        except (TypeError, ValueError):
            amount = 0.0
        fee_owed = round(pct * amount, 2)
        tx_data = {
            "date": check_row.get("check_date", ""),
            "amount": amount,
            "type": _coverage_to_type(check_row.get("coverage", "")),
            "fee_owed": fee_owed,
            "balance": 0,
            "deferred": 0,
            "recouped": 0,
            "fee_collected": 0,
            "reimbursed": 0,
            "total_collected": 0,
            "unpaid_payee_expense": 0,
            "outstanding_expense": 0,
            "ott": None,
            "notes": f"Check #{check_row.get('check_number', '')} — {check_row.get('coverage', '')}",
            "check_id": check_id,
        }
        db_module.create_transaction(claim_id, tx_data)

    flash("Check matched and transaction created.", "success")
    return redirect(url_for("checks_queue"))


def _coverage_to_type(coverage: str) -> str:
    c = (coverage or "").lower()
    if "draw" in c:
        return "Draw"
    if "escrow" in c:
        return "Escrow"
    return "Carrier"


# ---------------------------------------------------------------------------
# Assignees (claim-level split defaults) — HTMX fragments
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/assignees/new-row")
@login_required
def assignee_new_row(claim_id: int):
    claim = db_module.get_claim(claim_id)
    return render_template("partials/_assignee_edit.html",
                           claim=claim, assignee=None, roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())

@app.route("/claims/<int:claim_id>/assignees", methods=["POST"])
@login_required
def assignee_create(claim_id: int):
    data = _assignee_data_from_form()
    aid = db_module.create_assignee(claim_id, data)
    assignee = db_module.get_assignee(aid)
    claim = db_module.get_claim(claim_id)
    row_html = render_template("partials/_assignee_row.html",
                               claim=claim, assignee=assignee, roles=ASSIGNEE_ROLES)
    placeholder = '<tr id="assignee-new-row-placeholder"></tr>'
    total_html = _assignee_total_oob(claim_id)
    return row_html + placeholder + total_html

@app.route("/claims/<int:claim_id>/assignees/<int:aid>/edit")
@login_required
def assignee_edit_row(claim_id: int, aid: int):
    claim = db_module.get_claim(claim_id)
    assignee = db_module.get_assignee(aid)
    if not assignee:
        abort(404)
    return render_template("partials/_assignee_edit.html",
                           claim=claim, assignee=assignee, roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())

@app.route("/claims/<int:claim_id>/assignees/<int:aid>", methods=["PUT"])
@login_required
def assignee_update(claim_id: int, aid: int):
    data = _assignee_data_from_form()
    db_module.update_assignee(aid, data)
    assignee = db_module.get_assignee(aid)
    claim = db_module.get_claim(claim_id)
    row_html = render_template("partials/_assignee_row.html",
                               claim=claim, assignee=assignee, roles=ASSIGNEE_ROLES)
    return row_html + _assignee_total_oob(claim_id)

@app.route("/claims/<int:claim_id>/assignees/<int:aid>", methods=["DELETE"])
@login_required
def assignee_delete(claim_id: int, aid: int):
    db_module.delete_assignee(aid)
    return _assignee_total_oob(claim_id)

@app.route("/claims/<int:claim_id>/assignees/<int:aid>")
@login_required
def assignee_view_row(claim_id: int, aid: int):
    claim = db_module.get_claim(claim_id)
    assignee = db_module.get_assignee(aid)
    if not assignee:
        abort(404)
    return render_template("partials/_assignee_row.html",
                           claim=claim, assignee=assignee, roles=ASSIGNEE_ROLES)

def _recipient_name_from_id(recipient_id) -> str:
    """Look up a fee recipient's name by id; return empty string if not found."""
    if not recipient_id:
        return ""
    rec = db_module.get_fee_recipient(int(recipient_id))
    return rec["name"] if rec else ""

def _assignee_data_from_form() -> dict:
    rid = request.form.get("recipient_id") or None
    return {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "sort_order":   request.form.get("sort_order", 0),
        "recipient_id": rid,
    }

def _assignee_total_oob(claim_id: int) -> str:
    total = db_module.get_assignee_total_pct(claim_id)
    return f'<span id="assignee-total" hx-swap-oob="true">{total:,.2f}%</span>'


# ---------------------------------------------------------------------------
# Transaction detail page (full page)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>")
@login_required
def tx_detail(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    if not claim or not tx:
        abort(404)
    coverages = db_module.get_coverages(tx_id)
    disbursements = db_module.get_disbursements(tx_id)
    disburse_totals = _effective_split_basis(tx)
    has_custom_recouped = db_module.has_custom_recouped_splits(tx_id)
    include_recouped = not has_custom_recouped
    splits = [db_module.calc_split_amounts(s, disburse_totals, include_recouped=include_recouped)
              for s in db_module.get_splits(tx_id)]
    recouped_splits = db_module.get_recouped_splits(tx_id) if has_custom_recouped else []
    vendors = db_module.get_all_vendors()

    escrow_checklist = []
    if (tx.get("type") or "").lower() == "draw":
        escrow_checklist = db_module.get_escrow_checklist(claim_id)

    invoice_payments = []
    invoice_deferred_links = []
    claim_txs_for_select = []
    if (tx.get("type") or "").lower() == "fee invoice":
        invoice_payments = db_module.get_invoice_payments(tx_id)
        invoice_deferred_links = db_module.get_invoice_deferred_links(tx_id)
        claim_txs_for_select = db_module.get_claim_transactions_for_select(claim_id, exclude_tx_id=tx_id)

    recouped_links = db_module.get_recouped_deferred_links(tx_id)
    claim_txs_with_deferred = db_module.get_claim_txs_with_deferred(claim_id)

    # Outstanding deferred (deferred − recouped) — matches the summary card value
    _summary = db_module.get_transaction_summary(claim_id)
    claim_total_deferred = max(0.0, float(_summary.get("deferred_vs_recouped") or 0))

    tx_status = _tx_display_status(disbursements)
    return render_template(
        "transaction.html",
        claim=claim,
        tx=tx,
        splits=splits,
        coverages=coverages,
        disbursements=disbursements,
        disburse_totals=disburse_totals,
        has_custom_recouped=has_custom_recouped,
        recouped_splits=recouped_splits,
        vendors=vendors,
        escrow_checklist=escrow_checklist,
        invoice_payments=invoice_payments,
        invoice_deferred_links=invoice_deferred_links,
        claim_txs_for_select=claim_txs_for_select,
        recouped_links=recouped_links,
        claim_txs_with_deferred=claim_txs_with_deferred,
        roles=ASSIGNEE_ROLES,
        coverage_types=db_module.get_claim_coverage_types(claim_id),
        tx_types=TRANSACTION_TYPES,
        tx_status=tx_status,
        tx_status_css=_TX_STATUS_CSS[tx_status],
        tx_status_label=_TX_STATUS_LABEL[tx_status],
        claim_total_deferred=claim_total_deferred,
    )


# Transaction info (header fields) — view + HTMX inline edit

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/info")
@login_required
def tx_info_view(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    if not claim or not tx:
        abort(404)
    return render_template("partials/_tx_info.html", claim=claim, tx=tx,
                           tx_types=TRANSACTION_TYPES,
                           vendors=db_module.get_all_vendors(),
                           fee_invoices=db_module.get_fee_invoices(claim_id))


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/info/edit")
@login_required
def tx_info_edit(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    if not claim or not tx:
        abort(404)
    return render_template("partials/_tx_info_edit.html",
                           claim=claim, tx=tx, tx_types=TRANSACTION_TYPES,
                           vendors=db_module.get_all_vendors(),
                           fee_invoices=db_module.get_fee_invoices(claim_id))

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/info", methods=["POST"])
@login_required
def tx_info_save(claim_id: int, tx_id: int):
    tx = db_module.get_transaction(tx_id)
    if not tx:
        abort(404)
    data = dict(tx)
    data.update({
        "check_number":    request.form.get("check_number", "").strip(),
        "received_date":   _clean_date(request.form.get("received_date")),
        "payer":           request.form.get("payer", "").strip(),
        "payees_text":     request.form.get("payees_text", "").strip(),
        "endorsed":        bool(request.form.get("endorsed")),
        "void":            bool(request.form.get("void")),
        "type":            request.form.get("type", tx.get("type", "")),
        "date":            _clean_date(request.form.get("date")) or tx.get("date", ""),
        "notes":           request.form.get("notes", tx.get("notes", "")),
        "linked_escrow_id": request.form.get("linked_escrow_id") or None,
        "invoice_number":  request.form.get("invoice_number", "").strip(),
        "inv_for":         request.form.get("inv_for", "insured"),
        "inv_vendor_id":   request.form.get("inv_vendor_id") or None,
        "proc_method":         request.form.get("proc_method") or None,
        "proc_fee_invoice_id": request.form.get("proc_fee_invoice_id") or None,
    })
    db_module.update_transaction(tx_id, data)
    tx = db_module.get_transaction(tx_id)
    claim = db_module.get_claim(claim_id)
    info_html = render_template("partials/_tx_info.html", claim=claim, tx=tx,
                                tx_types=TRANSACTION_TYPES,
                                vendors=db_module.get_all_vendors(),
                                fee_invoices=db_module.get_fee_invoices(claim_id))
    return info_html + _check_status_oob(tx_id)


# ---------------------------------------------------------------------------
# Transaction Splits — HTMX fragments (on detail page)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/splits/new-row")
@login_required
def split_new_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    return render_template("partials/_split_edit.html",
                           claim=claim, tx=tx, split=None, roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/splits", methods=["POST"])
@login_required
def split_create(claim_id: int, tx_id: int):
    data = _split_data_from_form()
    sid = db_module.create_split(tx_id, data)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    basis = _effective_split_basis(tx) if tx else db_module.get_disbursement_totals(tx_id)
    include_recouped = not db_module.has_custom_recouped_splits(tx_id)
    split = db_module.calc_split_amounts(db_module.get_split(sid), basis, include_recouped=include_recouped)
    row_html = render_template("partials/_split_row.html",
                               claim=claim, tx=tx, split=split, roles=ASSIGNEE_ROLES)
    placeholder = '<tr id="split-new-row-placeholder"></tr>'
    total_html = _split_totals_only_oob(tx_id)
    return row_html + placeholder + total_html

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/splits/<int:sid>/edit")
@login_required
def split_edit_row(claim_id: int, tx_id: int, sid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    split = db_module.get_split(sid)
    if not split:
        abort(404)
    return render_template("partials/_split_edit.html",
                           claim=claim, tx=tx, split=split, roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/splits/<int:sid>", methods=["PUT"])
@login_required
def split_update(claim_id: int, tx_id: int, sid: int):
    data = _split_data_from_form()
    db_module.update_split(sid, data)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    basis = _effective_split_basis(tx) if tx else db_module.get_disbursement_totals(tx_id)
    include_recouped = not db_module.has_custom_recouped_splits(tx_id)
    split = db_module.calc_split_amounts(db_module.get_split(sid), basis, include_recouped=include_recouped)
    row_html = render_template("partials/_split_row.html",
                               claim=claim, tx=tx, split=split, roles=ASSIGNEE_ROLES)
    return row_html + _split_totals_only_oob(tx_id)

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/splits/<int:sid>", methods=["DELETE"])
@login_required
def split_delete(claim_id: int, tx_id: int, sid: int):
    db_module.delete_split(sid)
    return _split_totals_only_oob(tx_id)

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/splits/seed", methods=["POST"])
@login_required
def split_seed(claim_id: int, tx_id: int):
    """Re-clone claim assignees into this transaction's splits (replaces existing)."""
    db_module.get_db().execute(
        "DELETE FROM transaction_splits WHERE transaction_id = ?", (tx_id,))
    db_module.get_db().commit()
    db_module.clone_splits_from_assignees(claim_id, tx_id)
    tx = db_module.get_transaction(tx_id)
    claim = db_module.get_claim(claim_id)
    basis = _effective_split_basis(tx) if tx else db_module.get_disbursement_totals(tx_id)
    has_custom_recouped = db_module.has_custom_recouped_splits(tx_id)
    include_recouped = not has_custom_recouped
    splits = [db_module.calc_split_amounts(s, basis, include_recouped=include_recouped)
              for s in db_module.get_splits(tx_id)]
    disburse_totals = basis
    return render_template("partials/_splits_section.html",
                           claim=claim, tx=tx, splits=splits,
                           disburse_totals=disburse_totals, roles=ASSIGNEE_ROLES,
                           has_custom_recouped=has_custom_recouped)

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/splits/<int:sid>")
@login_required
def split_view_row(claim_id: int, tx_id: int, sid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    split = db_module.get_split(sid)
    if not split:
        abort(404)
    basis = _effective_split_basis(tx) if tx else {}
    include_recouped = not db_module.has_custom_recouped_splits(tx_id)
    split = db_module.calc_split_amounts(split, basis, include_recouped=include_recouped)
    return render_template("partials/_split_row.html",
                           claim=claim, tx=tx, split=split, roles=ASSIGNEE_ROLES)

def _effective_split_basis(tx: dict) -> dict:
    """Return the fee totals dict to use for split amount calculations.

    For Fee Invoice, the entire invoice amount is the fee — there are no
    disbursements, so the normal disbursement totals would all be zero.
    """
    if (tx.get("type") or "").lower() == "fee invoice":
        amount = float(tx.get("amount") or 0)
        return {
            "fee_owed": amount,
            "fee_deferred": 0.0,
            "fee_recouped": 0.0,
            "net_to_insured": 0.0,
            "to_vendors": 0.0,
            "total_disbursed": 0.0,
        }
    return db_module.get_disbursement_totals(tx["id"])


def _split_data_from_form() -> dict:
    rid = request.form.get("recipient_id") or None
    return {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "recipient_id": rid,
    }

def _split_totals_only_oob(tx_id: int) -> str:
    """Update just the split summary spans — no tbody replacement, no row flicker."""
    tx     = db_module.get_transaction(tx_id)
    basis  = _effective_split_basis(tx) if tx else db_module.get_disbursement_totals(tx_id)
    include_recouped = not db_module.has_custom_recouped_splits(tx_id)
    splits = [db_module.calc_split_amounts(s, basis, include_recouped=include_recouped)
              for s in db_module.get_splits(tx_id)]
    total_pct = sum(s.get("split_pct") or 0 for s in splits)
    total_fee = sum(s.get("fee_amount") or 0 for s in splits)
    css = "color:var(--red)" if abs(total_pct - 100) > 0.01 and splits else ""
    return (
        f'<span id="split-total-pct" hx-swap-oob="true" style="{css}">{total_pct:,.2f}%</span>'
        f'<span id="split-total-fee" hx-swap-oob="true">${total_fee:,.2f}</span>'
    )

def _split_rows_oob(tx_id: int) -> str:
    """Emit an OOB <tr> swap for each existing split row to refresh per-row amounts.

    Called when disbursement totals change (fee_recouped, fee_deferred, etc.) so that
    individual split rows reflect the updated basis without replacing the whole tbody.
    """
    tx = db_module.get_transaction(tx_id)
    if not tx:
        return ""
    claim = db_module.get_claim(tx["claim_id"])
    if not claim:
        return ""
    basis  = _effective_split_basis(tx)
    include_recouped = not db_module.has_custom_recouped_splits(tx_id)
    splits = [db_module.calc_split_amounts(s, basis, include_recouped=include_recouped)
              for s in db_module.get_splits(tx_id)]
    rows = ""
    for split in splits:
        row = render_template("partials/_split_row.html",
                              claim=claim, tx=tx, split=split, roles=ASSIGNEE_ROLES)
        # Add hx-swap-oob="true" so HTMX matches by id and swaps outerHTML
        rows += row.replace("<tr ", '<tr hx-swap-oob="true" ', 1)
    if not rows:
        return ""
    # Wrap in a hidden table so the browser parses <tr> in a valid table context.
    # Without this, bare <tr hx-swap-oob> elements get stripped by the HTML parser
    # and their content is injected as stray text into the primary swap target.
    return f'<table style="display:none"><tbody>{rows}</tbody></table>'

def _split_total_oob(tx_id: int) -> str:
    """Full tbody replacement — only used by seed (replaces all rows at once)."""
    tx    = db_module.get_transaction(tx_id)
    basis = _effective_split_basis(tx) if tx else db_module.get_disbursement_totals(tx_id)
    splits = [db_module.calc_split_amounts(s, basis) for s in db_module.get_splits(tx_id)]
    total_pct = sum(s.get("split_pct") or 0 for s in splits)
    total_fee = sum(s.get("fee_amount") or 0 for s in splits)
    css = "color:var(--red)" if abs(total_pct - 100) > 0.01 and splits else ""
    claim = db_module.get_claim(tx["claim_id"]) if tx else None
    rows_html = ""
    if claim and tx:
        for split in splits:
            rows_html += render_template("partials/_split_row.html",
                                         claim=claim, tx=tx, split=split,
                                         roles=ASSIGNEE_ROLES)
    tbody_oob = (
        f'<tbody id="split-tbody" hx-swap-oob="true">'
        f'{rows_html}'
        f'<tr id="split-new-row-placeholder"></tr>'
        f'</tbody>'
    )
    return (
        f'<span id="split-total-pct" hx-swap-oob="true" style="{css}">{total_pct:,.2f}%</span>'
        f'<span id="split-total-fee" hx-swap-oob="true">${total_fee:,.2f}</span>'
        + tbody_oob
    )


# ---------------------------------------------------------------------------
# Transaction Coverage — HTMX fragments (on detail page)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/coverage/new-row")
@login_required
def coverage_new_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    return render_template("partials/_coverage_edit.html",
                           claim=claim, tx=tx, cov=None,
                           coverage_types=db_module.get_claim_coverage_types(claim_id))

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/coverage", methods=["POST"])
@login_required
def coverage_create(claim_id: int, tx_id: int):
    data = _coverage_data_from_form()
    allowed = db_module.get_claim_coverage_types(claim_id)
    if data["coverage_type"] not in allowed:
        abort(400)
    db_module.ensure_coverage_type(data["coverage_type"])
    cid = db_module.create_coverage(tx_id, data)
    cov = db_module.get_coverage(cid)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    row_html = render_template("partials/_coverage_row.html",
                               claim=claim, tx=tx, cov=cov,
                               coverage_types=db_module.get_claim_coverage_types(claim_id))
    placeholder = '<tr id="coverage-new-row-placeholder"></tr>'
    return row_html + placeholder + _coverage_total_oob(tx_id)

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/coverage/<int:cid>/edit")
@login_required
def coverage_edit_row(claim_id: int, tx_id: int, cid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    cov = db_module.get_coverage(cid)
    if not cov:
        abort(404)
    return render_template("partials/_coverage_edit.html",
                           claim=claim, tx=tx, cov=cov,
                           coverage_types=db_module.get_claim_coverage_types(claim_id))

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/coverage/<int:cid>", methods=["PUT"])
@login_required
def coverage_update(claim_id: int, tx_id: int, cid: int):
    data = _coverage_data_from_form()
    allowed = db_module.get_claim_coverage_types(claim_id)
    if data["coverage_type"] not in allowed:
        abort(400)
    db_module.ensure_coverage_type(data["coverage_type"])
    db_module.update_coverage(cid, data)
    cov = db_module.get_coverage(cid)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    row_html = render_template("partials/_coverage_row.html",
                               claim=claim, tx=tx, cov=cov,
                               coverage_types=db_module.get_claim_coverage_types(claim_id))
    return row_html + _coverage_total_oob(tx_id)

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/coverage/<int:cid>", methods=["DELETE"])
@login_required
def coverage_delete(claim_id: int, tx_id: int, cid: int):
    db_module.delete_coverage(cid)
    return _coverage_total_oob(tx_id)

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/coverage/<int:cid>")
@login_required
def coverage_view_row(claim_id: int, tx_id: int, cid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    cov = db_module.get_coverage(cid)
    if not cov:
        abort(404)
    return render_template("partials/_coverage_row.html",
                           claim=claim, tx=tx, cov=cov,
                           coverage_types=db_module.get_claim_coverage_types(claim_id))

def _coverage_data_from_form() -> dict:
    return {
        "coverage_type": request.form.get("coverage_type", "").strip(),
        "amount": db_module._f(request.form.get("amount")),
    }

def _coverage_total_oob(tx_id: int) -> str:
    tx    = db_module.get_transaction(tx_id)
    covs  = db_module.get_coverages(tx_id)
    total = sum(c.get("amount") or 0 for c in covs)
    tx_amt = (tx.get("amount") or 0) if tx else 0
    pct    = total / tx_amt * 100 if tx_amt else 0
    pct_class = "text-green" if 99.99 <= pct <= 100.01 else "text-red"
    total_span = f'<span id="coverage-total" hx-swap-oob="true">${total:,.2f}</span>'
    pct_span   = f'<span id="coverage-pct" hx-swap-oob="true" class="{pct_class}">{pct:.1f}%</span>'
    return total_span + pct_span


# ---------------------------------------------------------------------------
# Disbursements — HTMX fragments (on transaction detail page)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/new-row")
@login_required
def disbursement_new_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    vendors = db_module.get_all_vendors()
    recipient_type = request.args.get("type", "insured")
    return render_template("partials/_disbursement_edit.html",
                           claim=claim, tx=tx, d=None,
                           recipient_type=recipient_type,
                           vendors=vendors)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements", methods=["POST"])
@login_required
def disbursement_create(claim_id: int, tx_id: int):
    data = _disbursement_data_from_form()
    did = db_module.create_disbursement(tx_id, data)
    d = db_module.get_disbursement(did)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    vendors = db_module.get_all_vendors()
    row_html = render_template("partials/_disbursement_row.html",
                               claim=claim, tx=tx, d=d, vendors=vendors)
    placeholder = '<tr id="disburse-new-row-placeholder"></tr>'
    totals_html = _disburse_totals_oob(tx_id, claim)
    return row_html + placeholder + totals_html


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/edit")
@login_required
def disbursement_edit_row(claim_id: int, tx_id: int, did: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    d = db_module.get_disbursement(did)
    if not d:
        abort(404)
    vendors = db_module.get_all_vendors()
    return render_template("partials/_disbursement_edit.html",
                           claim=claim, tx=tx, d=d,
                           recipient_type=d.get("recipient_type", "insured"),
                           vendors=vendors)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>",
           methods=["PUT"])
@login_required
def disbursement_update(claim_id: int, tx_id: int, did: int):
    data = _disbursement_data_from_form()
    db_module.update_disbursement(did, data)
    d = db_module.get_disbursement(did)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    vendors = db_module.get_all_vendors()
    row_html = render_template("partials/_disbursement_row.html",
                               claim=claim, tx=tx, d=d, vendors=vendors)
    return row_html + _disburse_totals_oob(tx_id, claim)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>",
           methods=["DELETE"])
@login_required
def disbursement_delete(claim_id: int, tx_id: int, did: int):
    db_module.delete_disbursement(did)
    claim = db_module.get_claim(claim_id)
    # Also delete the splits panel placeholder row that lives alongside this disbursement row
    panel_cleanup = f'<tr id="disburse-{did}-splits-panel" hx-swap-oob="delete"></tr>'
    return _disburse_totals_oob(tx_id, claim) + panel_cleanup


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>")
@login_required
def disbursement_view_row(claim_id: int, tx_id: int, did: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    d = db_module.get_disbursement(did)
    if not d:
        abort(404)
    vendors = db_module.get_all_vendors()
    return render_template("partials/_disbursement_row.html",
                           claim=claim, tx=tx, d=d, vendors=vendors)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/toggle-client",
           methods=["POST"])
@login_required
def disbursement_toggle_client(claim_id: int, tx_id: int, did: int):
    """Toggle whether the client/recipient has received their funds."""
    d = db_module.get_disbursement(did)
    if not d:
        abort(404)
    data = dict(d)
    data["client_received"] = 0 if d.get("client_received") else 1
    db_module.update_disbursement(did, data)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    vendors = db_module.get_all_vendors()
    d = db_module.get_disbursement(did)
    row_html = render_template("partials/_disbursement_row.html",
                               claim=claim, tx=tx, d=d, vendors=vendors)
    return row_html + _disburse_totals_oob(tx_id, claim) + _check_status_oob(tx_id)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/toggle-fee",
           methods=["POST"])
@login_required
def disbursement_toggle_fee(claim_id: int, tx_id: int, did: int):
    """Toggle fee collected: set to fee_owed, or back to 0 if already collected."""
    d = db_module.get_disbursement(did)
    if not d:
        abort(404)
    data = dict(d)
    fee_collectable = (d.get("fee_owed") or 0) - (d.get("fee_deferred") or 0) + (d.get("fee_recouped") or 0)
    already_collected = fee_collectable > 0 and (d.get("fee_collected") or 0) >= fee_collectable - 0.005
    data["fee_collected"] = 0 if already_collected else fee_collectable
    # leave fee_deferred unchanged
    db_module.update_disbursement(did, data)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    vendors = db_module.get_all_vendors()
    d = db_module.get_disbursement(did)
    row_html = render_template("partials/_disbursement_row.html",
                               claim=claim, tx=tx, d=d, vendors=vendors)
    return row_html + _disburse_totals_oob(tx_id, claim) + _check_status_oob(tx_id)


def _parse_int_or_none(val) -> int | None:
    """Return int if val is a non-empty digit string, else None."""
    if val and str(val).strip().lstrip('-').isdigit():
        return int(val)
    return None


def _disbursement_data_from_form() -> dict:
    fee_applies_raw = request.form.get("fee_applies")
    fee_applies = fee_applies_raw is not None and fee_applies_raw != "0"
    fee_pct = db_module._fopt(request.form.get("fee_pct"))
    # If the effective pct is 0, fee cannot apply
    if fee_pct is not None and fee_pct == 0:
        fee_applies = False
    return {
        "sort_order":     request.form.get("sort_order", 0),
        "recipient_type": request.form.get("recipient_type", "insured"),
        "vendor_id":      _parse_int_or_none(request.form.get("vendor_id")),
        "recipient_name": request.form.get("recipient_name", "").strip(),
        "amount":         db_module._f(request.form.get("amount")),
        "fee_applies":    fee_applies,
        "fee_pct":        fee_pct,
        "fee_owed":       db_module._f(request.form.get("fee_owed")),
        "fee_collected":  db_module._f(request.form.get("fee_collected")),
        "fee_deferred":   db_module._f(request.form.get("fee_deferred")),
        "fee_recouped":   db_module._f(request.form.get("fee_recouped")),
        "use_check_splits": True,
        "notes":          request.form.get("notes", "").strip(),
    }


def _disburse_totals_oob(tx_id: int, claim: dict) -> str:
    t = db_module.get_disbursement_totals(tx_id)
    tx = db_module.get_transaction(tx_id)
    tx_amount = (tx or {}).get("amount") or 0
    pct_html = ""
    remaining_html = ""
    if tx_amount:
        pct = (t["total_disbursed"] or 0) / tx_amount * 100
        if pct >= 99.99 and pct <= 100.01:
            pct_cls = "pct-exact"
        elif pct > 100.01:
            pct_cls = "pct-over"
        else:
            pct_cls = "pct-under"
        pct_html = f'<span id="disburse-total-pct" hx-swap-oob="true" class="disburse-pct {pct_cls}">{pct:.1f}%</span>'
        remaining = tx_amount - (t["total_disbursed"] or 0)
        if abs(remaining) < 0.005:
            rem_style = "color:var(--green)"
        elif remaining < 0:
            rem_style = "color:var(--red)"
        else:
            rem_style = ""
        remaining_html = f'<span id="disburse-remaining" hx-swap-oob="true" style="{rem_style}">${remaining:,.2f}</span>'
    return (
        f'<span id="disburse-total-disbursed" hx-swap-oob="true">${t["total_disbursed"]:,.2f}</span>'
        f'<span id="disburse-total-fee-owed" hx-swap-oob="true">${t["fee_owed"]:,.2f}</span>'
        f'<span id="disburse-total-fee-coll" hx-swap-oob="true">${t["fee_collected"]:,.2f}</span>'  # fee_collected is set by toggle to fee_owed−deferred+recouped
        f'<span id="disburse-total-deferred" hx-swap-oob="true">${t["fee_deferred"]:,.2f}</span>'
        f'<span id="disburse-sidebar-net" hx-swap-oob="true">${t["net_to_insured"]:,.2f}</span>'
        f'<span id="disburse-sidebar-vendors" hx-swap-oob="true">${t["to_vendors"]:,.2f}</span>'
        f'<span id="disburse-sidebar-fee-owed" hx-swap-oob="true">${t["fee_owed"]:,.2f}</span>'
        f'<span id="disburse-sidebar-fee-coll" hx-swap-oob="true">${max(0.0, t["fee_owed"] - t["fee_deferred"] + t["fee_recouped"]):,.2f}</span>'
        f'<span id="disburse-sidebar-deferred" hx-swap-oob="true">${t["fee_deferred"]:,.2f}</span>'
        f'<span id="disburse-sidebar-recouped" hx-swap-oob="true">${t["fee_recouped"]:,.2f}</span>'
        + pct_html
        + remaining_html
        + _split_totals_only_oob(tx_id)
        + _split_rows_oob(tx_id)
    )


# ---------------------------------------------------------------------------
# Disbursement Splits — HTMX fragments (per-disbursement custom payroll splits)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits-panel")
@login_required
def disburse_splits_panel(claim_id: int, tx_id: int, did: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    d = db_module.get_disbursement(did)
    if not d:
        abort(404)
    splits = db_module.get_disbursement_splits(did)
    return render_template("partials/_disburse_splits_panel.html",
                           claim=claim, tx=tx, d=d, splits=splits,
                           roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits/activate",
           methods=["POST"])
@login_required
def disburse_splits_activate(claim_id: int, tx_id: int, did: int):
    db_module.seed_disbursement_splits_from_tx(did, tx_id)
    d = db_module.get_disbursement(did)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    splits = db_module.get_disbursement_splits(did)
    panel = render_template("partials/_disburse_splits_panel.html",
                            claim=claim, tx=tx, d=d, splits=splits,
                            roles=ASSIGNEE_ROLES,
                            fee_recipients=db_module.get_all_fee_recipients())
    return panel + _disburse_split_btn_oob(d, claim, tx)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits/reset",
           methods=["POST"])
@login_required
def disburse_splits_reset(claim_id: int, tx_id: int, did: int):
    db_module.reset_disbursement_splits(did)
    d = db_module.get_disbursement(did)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    panel = render_template("partials/_disburse_splits_panel.html",
                            claim=claim, tx=tx, d=d, splits=[],
                            roles=ASSIGNEE_ROLES,
                            fee_recipients=db_module.get_all_fee_recipients())
    return panel + _disburse_split_btn_oob(d, claim, tx)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits/new-row")
@login_required
def disburse_split_new_row(claim_id: int, tx_id: int, did: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    d = db_module.get_disbursement(did)
    return render_template("partials/_disburse_split_edit.html",
                           claim=claim, tx=tx, d=d, split=None,
                           roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits",
           methods=["POST"])
@login_required
def disburse_split_create(claim_id: int, tx_id: int, did: int):
    rid = request.form.get("recipient_id") or None
    data = {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "recipient_id": rid,
    }
    sid = db_module.create_disbursement_split(did, data)
    d = db_module.get_disbursement(did)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    split = db_module.get_disbursement_split(sid)
    row_html = render_template("partials/_disburse_split_row.html",
                               claim=claim, tx=tx, d=d, split=split)
    placeholder = f'<tr id="dsplit-new-row-{did}"></tr>'
    return row_html + placeholder + _disburse_split_totals_oob(did)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits/<int:sid>/edit")
@login_required
def disburse_split_edit_row(claim_id: int, tx_id: int, did: int, sid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    d = db_module.get_disbursement(did)
    split = db_module.get_disbursement_split(sid)
    if not split:
        abort(404)
    return render_template("partials/_disburse_split_edit.html",
                           claim=claim, tx=tx, d=d, split=split,
                           roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits/<int:sid>",
           methods=["PUT"])
@login_required
def disburse_split_update(claim_id: int, tx_id: int, did: int, sid: int):
    rid = request.form.get("recipient_id") or None
    data = {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "recipient_id": rid,
    }
    db_module.update_disbursement_split(sid, data)
    d = db_module.get_disbursement(did)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    split = db_module.get_disbursement_split(sid)
    row_html = render_template("partials/_disburse_split_row.html",
                               claim=claim, tx=tx, d=d, split=split)
    return row_html + _disburse_split_totals_oob(did)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits/<int:sid>",
           methods=["DELETE"])
@login_required
def disburse_split_delete(claim_id: int, tx_id: int, did: int, sid: int):
    db_module.delete_disbursement_split(sid)
    return _disburse_split_totals_oob(did)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/disbursements/<int:did>/splits/<int:sid>")
@login_required
def disburse_split_view_row(claim_id: int, tx_id: int, did: int, sid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    d = db_module.get_disbursement(did)
    split = db_module.get_disbursement_split(sid)
    if not split:
        abort(404)
    return render_template("partials/_disburse_split_row.html",
                           claim=claim, tx=tx, d=d, split=split)


def _disburse_split_totals_oob(did: int) -> str:
    """OOB span updates for the disbursement splits mini-table footer."""
    splits = db_module.get_disbursement_splits(did)
    d = db_module.get_disbursement(did)
    fee_owed = (d.get("fee_owed") or 0) if d else 0
    total_pct = sum(s.get("split_pct") or 0 for s in splits)
    total_amt = round(fee_owed * total_pct / 100, 2)
    css = "color:var(--red)" if splits and abs(total_pct - 100) > 0.01 else ""
    return (
        f'<span id="dsplit-total-pct-{did}" hx-swap-oob="true" style="{css}">{total_pct:,.2f}%</span>'
        f'<span id="dsplit-total-amt-{did}" hx-swap-oob="true">${total_amt:,.2f}</span>'
    )


def _disburse_split_btn_oob(d: dict, claim: dict, tx: dict) -> str:
    """OOB update for the ⚙ icon button in the disbursement row after activate/reset."""
    did = d["id"]
    is_active = not d.get("use_check_splits", 1)
    btn_cls = "btn-icon btn-toggle-on" if is_active else "btn-icon"
    tooltip = "Custom splits active" if is_active else "Custom splits"
    panel_url = url_for("disburse_splits_panel", claim_id=claim["id"], tx_id=tx["id"], did=did)
    return (
        f'<button id="dsplit-btn-{did}" hx-swap-oob="true"'
        f' class="{btn_cls}"'
        f' hx-get="{panel_url}"'
        f' hx-target="#disburse-{did}-splits-panel"'
        f' hx-swap="outerHTML"'
        f' title="{tooltip}">⚙</button>'
    )


# ---------------------------------------------------------------------------
# Recouped Fee Custom Splits — HTMX fragments
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits/activate",
           methods=["POST"])
@login_required
def recouped_splits_activate(claim_id: int, tx_id: int):
    db_module.seed_recouped_splits_from_tx(tx_id)
    tx = db_module.get_transaction(tx_id)
    claim = db_module.get_claim(claim_id)
    basis = _effective_split_basis(tx) if tx else {}
    splits = [db_module.calc_split_amounts(s, basis, include_recouped=False)
              for s in db_module.get_splits(tx_id)]
    recouped_splits = db_module.get_recouped_splits(tx_id)
    section_html = render_template("partials/_splits_section.html",
                                   claim=claim, tx=tx, splits=splits,
                                   disburse_totals=basis, roles=ASSIGNEE_ROLES,
                                   has_custom_recouped=True)
    recouped_html = render_template("partials/_recouped_splits_section.html",
                                    claim=claim, tx=tx,
                                    recouped_splits=recouped_splits,
                                    disburse_totals=basis)
    oob_recouped = f'<div id="recouped-splits-section" hx-swap-oob="true">{recouped_html}</div>'
    return section_html + oob_recouped


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits/reset",
           methods=["POST"])
@login_required
def recouped_splits_reset(claim_id: int, tx_id: int):
    db_module.reset_recouped_splits(tx_id)
    tx = db_module.get_transaction(tx_id)
    claim = db_module.get_claim(claim_id)
    basis = _effective_split_basis(tx) if tx else {}
    splits = [db_module.calc_split_amounts(s, basis, include_recouped=True)
              for s in db_module.get_splits(tx_id)]
    section_html = render_template("partials/_splits_section.html",
                                   claim=claim, tx=tx, splits=splits,
                                   disburse_totals=basis, roles=ASSIGNEE_ROLES,
                                   has_custom_recouped=False)
    oob_recouped = '<div id="recouped-splits-section" hx-swap-oob="true"></div>'
    return section_html + oob_recouped


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits/new-row")
@login_required
def recouped_split_new_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    return render_template("partials/_recouped_split_edit.html",
                           claim=claim, tx=tx, split=None,
                           roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits",
           methods=["POST"])
@login_required
def recouped_split_create(claim_id: int, tx_id: int):
    rid = request.form.get("recipient_id") or None
    data = {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "recipient_id": rid,
    }
    sid = db_module.create_recouped_split(tx_id, data)
    split = db_module.get_recouped_split(sid)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    basis = _effective_split_basis(tx) if tx else {}
    row_html = render_template("partials/_recouped_split_row.html",
                               claim=claim, tx=tx, split=split, disburse_totals=basis)
    placeholder = '<tr id="recouped-split-new-row-placeholder"></tr>'
    return row_html + placeholder + _recouped_split_totals_oob(tx_id) + _split_totals_only_oob(tx_id)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits/<int:sid>/edit")
@login_required
def recouped_split_edit_row(claim_id: int, tx_id: int, sid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    split = db_module.get_recouped_split(sid)
    if not split:
        abort(404)
    return render_template("partials/_recouped_split_edit.html",
                           claim=claim, tx=tx, split=split,
                           roles=ASSIGNEE_ROLES,
                           fee_recipients=db_module.get_all_fee_recipients())


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits/<int:sid>",
           methods=["PUT"])
@login_required
def recouped_split_update(claim_id: int, tx_id: int, sid: int):
    rid = request.form.get("recipient_id") or None
    data = {
        "role":         request.form.get("role", "other"),
        "name":         _recipient_name_from_id(rid),
        "split_pct":    db_module._f(request.form.get("split_pct")),
        "recipient_id": rid,
    }
    db_module.update_recouped_split(sid, data)
    split = db_module.get_recouped_split(sid)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    basis = _effective_split_basis(tx) if tx else {}
    row_html = render_template("partials/_recouped_split_row.html",
                               claim=claim, tx=tx, split=split, disburse_totals=basis)
    return row_html + _recouped_split_totals_oob(tx_id) + _split_totals_only_oob(tx_id)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits/<int:sid>",
           methods=["DELETE"])
@login_required
def recouped_split_delete(claim_id: int, tx_id: int, sid: int):
    db_module.delete_recouped_split(sid)
    return _recouped_split_totals_oob(tx_id) + _split_totals_only_oob(tx_id)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-splits/<int:sid>")
@login_required
def recouped_split_view_row(claim_id: int, tx_id: int, sid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    split = db_module.get_recouped_split(sid)
    if not split:
        abort(404)
    basis = _effective_split_basis(tx) if tx else {}
    return render_template("partials/_recouped_split_row.html",
                           claim=claim, tx=tx, split=split, disburse_totals=basis)


def _recouped_split_totals_oob(tx_id: int) -> str:
    """OOB spans for the recouped splits section totals footer."""
    recouped_splits = db_module.get_recouped_splits(tx_id)
    tx = db_module.get_transaction(tx_id)
    basis = _effective_split_basis(tx) if tx else {}
    fee_recouped = basis.get("fee_recouped") or 0
    total_pct = sum(s.get("split_pct") or 0 for s in recouped_splits)
    total_amt = round(fee_recouped * total_pct / 100, 2)
    css = "color:var(--red)" if recouped_splits and abs(total_pct - 100) > 0.01 else ""
    return (
        f'<span id="recouped-split-total-pct" hx-swap-oob="true" style="{css}">{total_pct:,.2f}%</span>'
        f'<span id="recouped-split-total-fee" hx-swap-oob="true">${total_amt:,.2f}</span>'
    )


# ---------------------------------------------------------------------------
# Vendors — global management
# ---------------------------------------------------------------------------

@app.route("/vendors")
@login_required
def vendors_list():
    vendors = db_module.get_all_vendors()
    return render_template("vendors.html", vendors=vendors)


@app.route("/vendors/summary")
@login_required
def vendor_summary():
    data = db_module.get_global_vendor_activity()
    return render_template("vendor_summary.html", **data)


@app.route("/vendors/<int:vendor_id>/claims")
@login_required
def vendor_claims(vendor_id: int):
    vendor = db_module.get_vendor(vendor_id)
    if not vendor:
        abort(404)
    claims = db_module.get_vendor_claims(vendor_id)
    return render_template("vendor_claims.html", vendor=vendor, claims=claims)


@app.route("/claims/<int:claim_id>/vendors/<int:vendor_id>")
@login_required
def claim_vendor_detail(claim_id: int, vendor_id: int):
    claim  = db_module.get_claim(claim_id)
    vendor = db_module.get_vendor(vendor_id)
    if not claim or not vendor:
        abort(404)
    detail = db_module.get_claim_vendor_detail(claim_id, vendor_id)
    return render_template("vendor_claim_detail.html", claim=claim, vendor=vendor, **detail)


@app.route("/vendors/options")
@login_required
def vendors_options():
    vendors = db_module.get_all_vendors()
    return jsonify([{"value": str(v["id"]), "text": v["name"]} for v in vendors])


@app.route("/vendors/new-row")
@login_required
def vendor_new_row():
    return render_template("partials/_vendor_edit.html", v=None)


@app.route("/vendors/<int:vendor_id>/edit")
@login_required
def vendor_edit_row(vendor_id: int):
    v = db_module.get_vendor(vendor_id)
    if not v:
        abort(404)
    return render_template("partials/_vendor_edit.html", v=v)


@app.route("/vendors/<int:vendor_id>")
@login_required
def vendor_view_row(vendor_id: int):
    v = db_module.get_vendor(vendor_id)
    if not v:
        abort(404)
    return render_template("partials/_vendor_row.html", v=v)


@app.route("/vendors", methods=["POST"])
@login_required
def vendor_create():
    data = {
        "name":  request.form.get("name", "").strip(),
        "phone": request.form.get("phone", "").strip(),
        "email": request.form.get("email", "").strip(),
        "notes": request.form.get("notes", "").strip(),
    }
    if not data["name"]:
        return "Name required", 400
    try:
        vid = db_module.create_vendor(data)
    except Exception:
        return "Vendor name already exists", 409
    v = db_module.get_vendor(vid)
    # If called from Tom Select inline creation, return JSON
    if request.headers.get("Accept", "").startswith("application/json"):
        return jsonify({"value": str(v["id"]), "text": v["name"]})
    row_html = render_template("partials/_vendor_row.html", v=v)
    placeholder = '<tr id="vendor-new-row-placeholder"></tr>'
    return row_html + placeholder


@app.route("/vendors/<int:vendor_id>", methods=["PUT"])
@login_required
def vendor_update(vendor_id: int):
    data = {
        "name":  request.form.get("name", "").strip(),
        "phone": request.form.get("phone", "").strip(),
        "email": request.form.get("email", "").strip(),
        "notes": request.form.get("notes", "").strip(),
    }
    db_module.update_vendor(vendor_id, data)
    v = db_module.get_vendor(vendor_id)
    return render_template("partials/_vendor_row.html", v=v)


@app.route("/vendors/<int:vendor_id>", methods=["DELETE"])
@login_required
def vendor_delete(vendor_id: int):
    db_module.delete_vendor(vendor_id)
    return ""


# ---------------------------------------------------------------------------
# Fee Recipients — global management
# ---------------------------------------------------------------------------

@app.route("/fee-recipients")
@login_required
def fee_recipients_list():
    recipients = db_module.get_all_fee_recipients()
    return render_template("fee_recipients.html", recipients=recipients,
                           roles=ASSIGNEE_ROLES)


@app.route("/fee-recipients/options")
@login_required
def fee_recipients_options():
    recipients = db_module.get_all_fee_recipients()
    return jsonify([{"value": str(r["id"]), "text": r["name"]} for r in recipients])


@app.route("/fee-recipients", methods=["POST"])
@login_required
def fee_recipient_create():
    data = {
        "name":         request.form.get("name", "").strip(),
        "company_name": request.form.get("company_name", "").strip(),
        "default_role": request.form.get("default_role", "").strip(),
    }
    if not data["name"]:
        return "Name required", 400
    try:
        rid = db_module.create_fee_recipient(data)
    except Exception:
        return "Name already exists", 409
    recipients = db_module.get_all_fee_recipients()
    return render_template("fee_recipients.html", recipients=recipients,
                           roles=ASSIGNEE_ROLES)


@app.route("/fee-recipients/<int:rid>", methods=["PUT"])
@login_required
def fee_recipient_update(rid: int):
    data = {
        "name":            request.form.get("name", "").strip(),
        "company_name":    request.form.get("company_name", "").strip(),
        "default_role":    request.form.get("default_role", "").strip(),
        "payroll_visible": request.form.get("payroll_visible", "0"),
    }
    db_module.update_fee_recipient(rid, data)
    recipients = db_module.get_all_fee_recipients()
    return render_template("fee_recipients.html", recipients=recipients,
                           roles=ASSIGNEE_ROLES)


@app.route("/fee-recipients/<int:rid>", methods=["DELETE"])
@login_required
def fee_recipient_delete(rid: int):
    db_module.delete_fee_recipient(rid)
    recipients = db_module.get_all_fee_recipients()
    return render_template("fee_recipients.html", recipients=recipients,
                           roles=ASSIGNEE_ROLES)


# ---------------------------------------------------------------------------
# Invoice Payments — HTMX fragments (on transaction detail page)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/payments/new-row")
@login_required
def payment_new_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    claim_txs_for_select = db_module.get_claim_transactions_for_select(claim_id, exclude_tx_id=tx_id)
    return render_template("partials/_inv_pay_edit.html",
                           claim=claim, tx=tx,
                           p=None,
                           claim_txs_for_select=claim_txs_for_select)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/payments", methods=["POST"])
@login_required
def payment_create(claim_id: int, tx_id: int):
    data = {
        "date":        (_clean_date(request.form.get("date")) or ""),
        "amount":      db_module._f(request.form.get("amount")),
        "method":      request.form.get("method", "").strip(),
        "reference":   request.form.get("reference", "").strip(),
        "linked_tx_id": request.form.get("linked_tx_id") or None,
        "notes":       request.form.get("notes", "").strip(),
    }
    pid = db_module.create_invoice_payment(tx_id, data)
    p = db_module.get_invoice_payment(pid)
    # Re-fetch with joined fields
    payments = db_module.get_invoice_payments(tx_id)
    p_full = next((x for x in payments if x["id"] == pid), p)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    row_html = render_template("partials/_inv_pay_row.html", claim=claim, tx=tx, p=p_full)
    placeholder = '<tr id="inv-pay-new-row-placeholder"></tr>'
    return row_html + placeholder + _inv_amounts_oob(tx_id, tx)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/payments/<int:pid>/edit")
@login_required
def payment_edit_row(claim_id: int, tx_id: int, pid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    payments = db_module.get_invoice_payments(tx_id)
    p = next((x for x in payments if x["id"] == pid), None)
    if not p:
        abort(404)
    claim_txs_for_select = db_module.get_claim_transactions_for_select(claim_id, exclude_tx_id=tx_id)
    return render_template("partials/_inv_pay_edit.html",
                           claim=claim, tx=tx, p=p,
                           claim_txs_for_select=claim_txs_for_select)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/payments/<int:pid>/view")
@login_required
def payment_view_row(claim_id: int, tx_id: int, pid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    payments = db_module.get_invoice_payments(tx_id)
    p = next((x for x in payments if x["id"] == pid), None)
    if not p:
        abort(404)
    return render_template("partials/_inv_pay_row.html", claim=claim, tx=tx, p=p)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/payments/<int:pid>", methods=["PUT"])
@login_required
def payment_update(claim_id: int, tx_id: int, pid: int):
    data = {
        "date":         (_clean_date(request.form.get("date")) or ""),
        "amount":       db_module._f(request.form.get("amount")),
        "method":       request.form.get("method", "").strip(),
        "reference":    request.form.get("reference", "").strip(),
        "linked_tx_id": request.form.get("linked_tx_id") or None,
        "notes":        request.form.get("notes", "").strip(),
    }
    db_module.update_invoice_payment(pid, data)
    payments = db_module.get_invoice_payments(tx_id)
    p = next((x for x in payments if x["id"] == pid), None)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    row_html = render_template("partials/_inv_pay_row.html", claim=claim, tx=tx, p=p)
    return row_html + _inv_amounts_oob(tx_id, tx)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/payments/<int:pid>", methods=["DELETE"])
@login_required
def payment_delete(claim_id: int, tx_id: int, pid: int):
    db_module.delete_invoice_payment(pid)
    tx = db_module.get_transaction(tx_id)
    return _inv_amounts_oob(tx_id, tx)


def _inv_amounts_oob(tx_id: int, tx: dict) -> str:
    """OOB updates for payment section total + sidebar paid/balance (called on payment add/delete)."""
    payments = db_module.get_invoice_payments(tx_id)
    total_paid = sum((p.get("amount") or 0) for p in payments)
    invoice_amount = float(tx.get("amount") or 0) if tx else 0.0
    balance = invoice_amount - total_paid
    section_total = f'<span id="inv-pay-total" hx-swap-oob="true">{money_filter(total_paid)}</span>'
    sidebar_paid    = f'<span id="inv-sidebar-paid" hx-swap-oob="true">{money_filter(total_paid)}</span>'
    sidebar_balance = f'<span id="inv-sidebar-balance" hx-swap-oob="true">{money_filter(balance)}</span>'
    return section_total + sidebar_paid + sidebar_balance


def _check_toggle_area_html(claim_id: int, tx: dict) -> str:
    """Return the <div id="check-toggle-area"> HTML for outerHTML swap on toggle routes."""
    tx_id = tx["id"]
    client_on = bool(tx.get("inv_client_paid"))
    fee_on = bool(tx.get("inv_fee_received"))
    released = client_on and fee_on
    status_css = "status-released" if released else "status-inv-pending"
    status_label = "Released" if released else "Pending"
    client_cls = "btn-toggle btn-toggle-on" if client_on else "btn-toggle"
    client_icon = "✓" if client_on else "○"
    client_title = "Client paid — click to undo" if client_on else "Mark client as paid"
    fee_cls = "btn-toggle btn-toggle-on" if fee_on else "btn-toggle"
    fee_icon = "✓" if fee_on else "○"
    fee_title = "Fee received — click to undo" if fee_on else "Mark fee as received"
    toggle_url_client = url_for("toggle_inv_client", claim_id=claim_id, tx_id=tx_id)
    toggle_url_fee = url_for("toggle_inv_fee", claim_id=claim_id, tx_id=tx_id)
    return (
        f'<div id="check-toggle-area">'
        f'<div class="sidebar-value" style="font-size:14px">'
        f'<span id="tx-sidebar-status" class="status-badge {status_css}">{status_label}</span>'
        f'</div>'
        f'<div style="display:flex; gap:6px; flex-wrap:wrap; margin-top:6px">'
        f'<button class="{client_cls}"'
        f' hx-post="{toggle_url_client}"'
        f' hx-target="#check-toggle-area"'
        f' hx-swap="outerHTML"'
        f' title="{client_title}">'
        f'{client_icon} Client Paid'
        f'</button>'
        f'<button class="{fee_cls}"'
        f' hx-post="{toggle_url_fee}"'
        f' hx-target="#check-toggle-area"'
        f' hx-swap="outerHTML"'
        f' title="{fee_title}">'
        f'{fee_icon} Fee Received'
        f'</button>'
        f'</div>'
        f'</div>'
    )


def _check_row_status_oob(tx_id: int, tx: dict) -> str:
    """OOB update for the status badge in the claim detail transaction row."""
    released = bool(tx.get("inv_client_paid")) and bool(tx.get("inv_fee_received"))
    css = "status-released" if released else "status-inv-pending"
    label = "Released" if released else "Pending"
    return (f'<span id="tx-status-{tx_id}" hx-swap-oob="true" '
            f'class="status-badge {css}">{label}</span>')


def _tx_display_status(disbursements: list) -> str:
    """Return 'released', 'fee_pending', 'client_pending', or 'pending' for a tx."""
    if not disbursements:
        return "pending"
    all_client = all(d.get("client_received") for d in disbursements)
    fee_disburse = [d for d in disbursements
                    if d.get("fee_applies")
                    and (d.get("fee_owed") or 0) - (d.get("fee_deferred") or 0) > 0.005]
    all_fee = len(fee_disburse) == 0 or all(
        (d.get("fee_collected") or 0) >= (d.get("fee_owed") or 0) - (d.get("fee_deferred") or 0) - 0.005
        for d in fee_disburse
    )
    if all_client and all_fee:
        return "released"
    elif all_client:
        return "fee_pending"
    elif all_fee:
        return "client_pending"
    return "pending"


_TX_STATUS_CSS = {
    "released":       "status-released",
    "fee_pending":    "status-fee-pending",
    "client_pending": "status-client-pending",
    "pending":        "status-inv-pending",
}
_TX_STATUS_LABEL = {
    "released":       "Released",
    "fee_pending":    "Fee Pending",
    "client_pending": "Client Pending",
    "pending":        "Pending",
}


def _check_status_oob(tx_id: int) -> str:
    """OOB spans updating sidebar and tx-row status based on disbursement state."""
    tx = db_module.get_transaction(tx_id)
    if tx and tx.get("void"):
        css, label = "status-void", "Void"
    else:
        disbursements = db_module.get_disbursements(tx_id)
        status = _tx_display_status(disbursements)
        css = _TX_STATUS_CSS[status]
        label = _TX_STATUS_LABEL[status]
    return (
        f'<span id="tx-sidebar-status" hx-swap-oob="true" class="status-badge {css}">{label}</span>'
        f'<span id="tx-status-{tx_id}" hx-swap-oob="true" class="status-badge {css}">{label}</span>'
    )


# ---------------------------------------------------------------------------
# Check status toggles (carrier / draw / escrow endorsed)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/toggle-inv-client", methods=["POST"])
@login_required
def toggle_inv_client(claim_id: int, tx_id: int):
    tx = db_module.get_transaction(tx_id)
    if not tx:
        abort(404)
    data = dict(tx)
    data["inv_client_paid"] = 0 if tx.get("inv_client_paid") else 1
    db_module.update_transaction(tx_id, data)
    tx = db_module.get_transaction(tx_id)
    return _check_toggle_area_html(claim_id, tx) + _check_row_status_oob(tx_id, tx)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/toggle-inv-fee", methods=["POST"])
@login_required
def toggle_inv_fee(claim_id: int, tx_id: int):
    tx = db_module.get_transaction(tx_id)
    if not tx:
        abort(404)
    data = dict(tx)
    data["inv_fee_received"] = 0 if tx.get("inv_fee_received") else 1
    db_module.update_transaction(tx_id, data)
    tx = db_module.get_transaction(tx_id)
    return _check_toggle_area_html(claim_id, tx) + _check_row_status_oob(tx_id, tx)


# ---------------------------------------------------------------------------
# Invoice Deferred Links — HTMX fragments (on transaction detail page)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/deferred-links/new-row")
@login_required
def deferred_link_new_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    claim_txs_for_select = db_module.get_claim_transactions_for_select(claim_id, exclude_tx_id=tx_id)
    return render_template("partials/_deferred_link_edit.html",
                           claim=claim, tx=tx,
                           claim_txs_for_select=claim_txs_for_select)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/deferred-links", methods=["POST"])
@login_required
def deferred_link_create(claim_id: int, tx_id: int):
    deferred_tx_id = request.form.get("deferred_tx_id")
    if not deferred_tx_id:
        return "deferred_tx_id required", 400
    lid = db_module.create_invoice_deferred_link(tx_id, int(deferred_tx_id))
    links = db_module.get_invoice_deferred_links(tx_id)
    link = next((x for x in links if x["id"] == lid), None)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    row_html = render_template("partials/_deferred_link_row.html", claim=claim, tx=tx, link=link)
    placeholder = '<tr id="deferred-link-new-row-placeholder"></tr>'
    return row_html + placeholder


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/deferred-links/<int:lid>/edit")
@login_required
def deferred_link_edit_row(claim_id: int, tx_id: int, lid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    links = db_module.get_invoice_deferred_links(tx_id)
    link = next((l for l in links if l["id"] == lid), None)
    if not link:
        abort(404)
    claim_txs_for_select = db_module.get_claim_transactions_for_select(claim_id, exclude_tx_id=tx_id)
    return render_template("partials/_deferred_link_edit.html",
                           claim=claim, tx=tx, link=link,
                           claim_txs_for_select=claim_txs_for_select)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/deferred-links/<int:lid>/view")
@login_required
def deferred_link_view_row(claim_id: int, tx_id: int, lid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    links = db_module.get_invoice_deferred_links(tx_id)
    link = next((l for l in links if l["id"] == lid), None)
    if not link:
        abort(404)
    return render_template("partials/_deferred_link_row.html", claim=claim, tx=tx, link=link)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/deferred-links/<int:lid>",
           methods=["PUT"])
@login_required
def deferred_link_update(claim_id: int, tx_id: int, lid: int):
    deferred_tx_id = request.form.get("deferred_tx_id")
    if not deferred_tx_id:
        abort(400)
    db_module.update_invoice_deferred_link(lid, int(deferred_tx_id), tx_id)
    links = db_module.get_invoice_deferred_links(tx_id)
    link = next((l for l in links if l["id"] == lid), None)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    return render_template("partials/_deferred_link_row.html", claim=claim, tx=tx, link=link)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/deferred-links/<int:lid>",
           methods=["DELETE"])
@login_required
def deferred_link_delete(claim_id: int, tx_id: int, lid: int):
    db_module.delete_invoice_deferred_link(lid, tx_id)
    return ""


# ---------------------------------------------------------------------------
# Recouped Deferred Links — HTMX fragments (on transaction detail page)
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-links/new-row")
@login_required
def recouped_link_new_row(claim_id: int, tx_id: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    claim_txs_with_deferred = db_module.get_claim_txs_with_deferred(claim_id)
    return render_template("partials/_recouped_link_edit.html",
                           claim=claim, tx=tx,
                           claim_txs_with_deferred=claim_txs_with_deferred)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-links", methods=["POST"])
@login_required
def recouped_link_create(claim_id: int, tx_id: int):
    deferred_tx_id = request.form.get("deferred_tx_id")
    if not deferred_tx_id:
        return "deferred_tx_id required", 400
    lid = db_module.create_recouped_deferred_link(tx_id, int(deferred_tx_id))
    links = db_module.get_recouped_deferred_links(tx_id)
    link = next((x for x in links if x["id"] == lid), None)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    row_html = render_template("partials/_recouped_link_row.html", claim=claim, tx=tx, link=link)
    placeholder = '<tr id="recouped-link-new-row-placeholder"></tr>'
    return row_html + placeholder


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-links/<int:lid>/edit")
@login_required
def recouped_link_edit_row(claim_id: int, tx_id: int, lid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    links = db_module.get_recouped_deferred_links(tx_id)
    link = next((l for l in links if l["id"] == lid), None)
    if not link:
        abort(404)
    claim_txs_with_deferred = db_module.get_claim_txs_with_deferred(claim_id)
    return render_template("partials/_recouped_link_edit.html",
                           claim=claim, tx=tx, link=link,
                           claim_txs_with_deferred=claim_txs_with_deferred)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-links/<int:lid>/view")
@login_required
def recouped_link_view_row(claim_id: int, tx_id: int, lid: int):
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    links = db_module.get_recouped_deferred_links(tx_id)
    link = next((l for l in links if l["id"] == lid), None)
    if not link:
        abort(404)
    return render_template("partials/_recouped_link_row.html", claim=claim, tx=tx, link=link)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-links/<int:lid>",
           methods=["PUT"])
@login_required
def recouped_link_update(claim_id: int, tx_id: int, lid: int):
    deferred_tx_id = request.form.get("deferred_tx_id")
    if not deferred_tx_id:
        abort(400)
    db_module.update_recouped_deferred_link(lid, int(deferred_tx_id))
    links = db_module.get_recouped_deferred_links(tx_id)
    link = next((l for l in links if l["id"] == lid), None)
    claim = db_module.get_claim(claim_id)
    tx = db_module.get_transaction(tx_id)
    return render_template("partials/_recouped_link_row.html", claim=claim, tx=tx, link=link)


@app.route("/claims/<int:claim_id>/transactions/<int:tx_id>/recouped-links/<int:lid>",
           methods=["DELETE"])
@login_required
def recouped_link_delete(claim_id: int, tx_id: int, lid: int):
    db_module.delete_recouped_deferred_link(lid)
    return ""


# ---------------------------------------------------------------------------
# Payroll — claim payroll page
# ---------------------------------------------------------------------------

@app.route("/claims/<int:claim_id>/payroll")
@login_required
def claim_payroll(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim:
        abort(404)
    assignees = db_module.get_assignees(claim_id)
    payroll_data = db_module.get_claim_payroll_data(claim_id)
    historic = db_module.get_claim_historic_assignees(claim_id)
    current_names = {a["name"] for a in assignees}
    historic_names = {h["name"] for h in historic}
    has_historic = bool(historic_names - current_names) or _has_split_pct_change(historic, assignees)
    return render_template(
        "payroll.html",
        claim=claim,
        assignees=assignees,
        payroll_data=payroll_data,
        has_historic=has_historic,
        fee_recipients=db_module.get_all_fee_recipients(),
        roles=ASSIGNEE_ROLES,
    )


def _has_split_pct_change(historic: list[dict], assignees: list[dict]) -> bool:
    """True if any current assignee has a different split_pct than their historic min/max."""
    current_map = {a["name"]: a["split_pct"] for a in assignees}
    for h in historic:
        name = h["name"]
        if name in current_map:
            if abs(float(current_map[name] or 0) - float(h["split_pct"] or 0)) > 0.01:
                return True
    return False


@app.route("/claims/<int:claim_id>/payroll/historic-panel")
@login_required
def claim_payroll_historic(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim:
        abort(404)
    historic = db_module.get_claim_historic_assignees(claim_id)
    return render_template("partials/_payroll_historic_panel.html",
                           claim=claim, historic=historic)


@app.route("/claims/<int:claim_id>/client-view")
@login_required
def claim_client_view(claim_id: int):
    claim = db_module.get_claim(claim_id)
    if not claim:
        abort(404)
    data = db_module.get_client_view_data(claim_id)
    from datetime import date
    return render_template("client_view.html", claim=claim, today=date.today().strftime("%B %d, %Y"), **data)


# ---------------------------------------------------------------------------
# Payroll — site-wide payroll page
# ---------------------------------------------------------------------------

@app.route("/payroll")
@login_required
def site_payroll():
    today = date.today()
    default_from = "2000-01-01"   # wide default — user can narrow with the filter
    default_to = today.isoformat()
    date_from = request.args.get("date_from", default_from)
    date_to   = request.args.get("date_to",   default_to)
    payroll = db_module.get_site_payroll_data(date_from, date_to)
    return render_template(
        "site_payroll.html",
        payroll=payroll,
        date_from=date_from,
        date_to=date_to,
    )


@app.route("/payroll/export.csv")
@login_required
def site_payroll_export():
    today = date.today()
    date_from = request.args.get("date_from", "2000-01-01")
    date_to   = request.args.get("date_to",   today.isoformat())
    payroll = db_module.get_site_payroll_data(date_from, date_to)

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([
        "Name", "Role", "Section", "Job #", "Claim #", "Insured",
        "Check #", "Date", "Tx Type", "Split %",
        "Collected Amount", "Deferred Amount", "Paid", "Paid Date",
    ])
    for p in payroll["payees"]:
        for row in p["rows"]:
            writer.writerow([
                p["name"], row["role"], "Received",
                row["job_number"], row["claim_number"], row["insured_name"],
                row["check_number"], row["date"], row["type"],
                row["split_pct"], row["earned"], "",
                row["payroll_paid"], row["payroll_paid_date"],
            ])
        for row in p["pending_rows"]:
            writer.writerow([
                p["name"], row["role"], "Pending",
                row["job_number"], row["claim_number"], row["insured_name"],
                row["check_number"], row["date"], row["type"],
                row["split_pct"], "", row["deferred"],
                "", "",
            ])
        for row in p["exp_rows"]:
            writer.writerow([
                p["name"], row["role"], "Expense",
                row["job_number"], row["claim_number"], row["insured_name"],
                "", row["date"], "Expense",
                row["split_pct"], row["owed"], "",
                row["payroll_paid"], row["payroll_paid_date"],
            ])

    out.seek(0)
    filename = f"payroll_{date_from}_{date_to}.csv"
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/payroll/mark-paid", methods=["POST"])
@login_required
def site_payroll_mark_paid_bulk():
    paid_date = request.form.get("paid_date", "").strip() or date.today().isoformat()
    split_ids = [int(x) for x in request.form.getlist("split_ids[]") if x]
    exp_split_ids = [int(x) for x in request.form.getlist("exp_split_ids[]") if x]
    if split_ids:
        db_module.mark_splits_paid(split_ids, paid_date)
    if exp_split_ids:
        db_module.mark_expense_splits_paid(exp_split_ids, paid_date)
    # Refresh the page after bulk action
    date_from = request.form.get("date_from", "2000-01-01")
    date_to   = request.form.get("date_to",   date.today().isoformat())
    return redirect(url_for("site_payroll", date_from=date_from, date_to=date_to))


@app.route("/payroll/splits/<int:sid>/toggle-paid", methods=["POST"])
@login_required
def site_split_toggle_paid(sid: int):
    paid_date = request.form.get("paid_date", "").strip() or date.today().isoformat()
    db_module.mark_split_paid_toggle(sid, paid_date)
    # Return updated status badge for this split row
    split = db_module.get_split(sid)
    if not split:
        return "", 404
    earned = db_module._payroll_earned_for_split(split)
    paid = float(split.get("payroll_paid") or 0)
    status = db_module._payroll_status(earned, paid)
    badge = _payroll_badge_html(status, paid, split.get("payroll_paid_date") or "", sid=sid, kind="splits")
    return f'<span id="split-status-{sid}" hx-swap-oob="true">{badge}</span>'


@app.route("/payroll/expense-splits/<int:sid>/toggle-paid", methods=["POST"])
@login_required
def site_expense_split_toggle_paid(sid: int):
    paid_date = request.form.get("paid_date", "").strip() or date.today().isoformat()
    db_module.mark_expense_split_paid_toggle(sid, paid_date)
    es_row = db_module._row(db_module.get_db(),
                            "SELECT * FROM expense_splits WHERE id=?", (sid,))
    if not es_row:
        return "", 404
    es_row = dict(es_row)
    exp_row = db_module._row(db_module.get_db(),
                             "SELECT wp_amount FROM expenses WHERE id=?", (es_row["expense_id"],))
    wp_amount = float(exp_row["wp_amount"] or 0) if exp_row else 0.0
    earned = round(wp_amount * float(es_row.get("split_pct") or 0) / 100, 2)
    paid = float(es_row.get("payroll_paid") or 0)
    status = db_module._payroll_status(earned, paid)
    badge = _payroll_badge_html(status, paid, es_row.get("payroll_paid_date") or "", sid=sid, kind="expense-splits")
    return f'<span id="exp-split-status-{sid}" hx-swap-oob="true">{badge}</span>'


@app.route("/payroll/splits/<int:sid>/set-date", methods=["POST"])
@login_required
def site_split_set_date(sid: int):
    paid_date = request.form.get("paid_date", "").strip()
    db_module.get_db().execute(
        "UPDATE transaction_splits SET payroll_paid_date=? WHERE id=?",
        (paid_date or None, sid),
    )
    db_module.get_db().commit()
    split = db_module.get_split(sid)
    if not split:
        return "", 404
    earned = db_module._payroll_earned_for_split(split)
    paid = float(split.get("payroll_paid") or 0)
    status = db_module._payroll_status(earned, paid)
    badge = _payroll_badge_html(status, paid, split.get("payroll_paid_date") or "", sid=sid, kind="splits")
    return f'<span id="split-status-{sid}" hx-swap-oob="true">{badge}</span>'


@app.route("/payroll/expense-splits/<int:sid>/set-date", methods=["POST"])
@login_required
def site_expense_split_set_date(sid: int):
    paid_date = request.form.get("paid_date", "").strip()
    db_module.get_db().execute(
        "UPDATE expense_splits SET payroll_paid_date=? WHERE id=?",
        (paid_date or None, sid),
    )
    db_module.get_db().commit()
    es_row = db_module._row(db_module.get_db(), "SELECT * FROM expense_splits WHERE id=?", (sid,))
    if not es_row:
        return "", 404
    es_row = dict(es_row)
    exp_row = db_module._row(db_module.get_db(),
                             "SELECT wp_amount FROM expenses WHERE id=?", (es_row["expense_id"],))
    wp_amount = float(exp_row["wp_amount"] or 0) if exp_row else 0.0
    earned = round(wp_amount * float(es_row.get("split_pct") or 0) / 100, 2)
    paid = float(es_row.get("payroll_paid") or 0)
    status = db_module._payroll_status(earned, paid)
    badge = _payroll_badge_html(status, paid, es_row.get("payroll_paid_date") or "", sid=sid, kind="expense-splits")
    return f'<span id="exp-split-status-{sid}" hx-swap-oob="true">{badge}</span>'


@app.route("/payroll/splits/<int:sid>/set-paid", methods=["POST"])
@login_required
def site_split_set_paid(sid: int):
    try:
        amount = float(request.form.get("amount", 0))
    except (TypeError, ValueError):
        amount = 0.0
    paid_date = request.form.get("paid_date", "").strip() or None
    db_module.set_split_paid_amount(sid, amount)
    db_module.get_db().execute(
        "UPDATE transaction_splits SET payroll_paid_date=? WHERE id=?", (paid_date, sid)
    )
    db_module.get_db().commit()
    split = db_module.get_split(sid)
    if not split:
        return "", 404
    earned = db_module._payroll_earned_for_split(split)
    paid = float(split.get("payroll_paid") or 0)
    status = db_module._payroll_status(earned, paid)
    badge = _payroll_badge_html(status, paid, split.get("payroll_paid_date") or "", sid=sid, kind="splits")
    return f'<span id="split-status-{sid}">{badge}</span>'


@app.route("/payroll/expense-splits/<int:sid>/set-paid", methods=["POST"])
@login_required
def site_expense_split_set_paid(sid: int):
    try:
        amount = float(request.form.get("amount", 0))
    except (TypeError, ValueError):
        amount = 0.0
    paid_date = request.form.get("paid_date", "").strip() or None
    db_module.set_expense_split_paid_amount(sid, amount)
    db_module.get_db().execute(
        "UPDATE expense_splits SET payroll_paid_date=? WHERE id=?", (paid_date, sid)
    )
    db_module.get_db().commit()
    es_row = db_module._row(db_module.get_db(), "SELECT * FROM expense_splits WHERE id=?", (sid,))
    if not es_row:
        return "", 404
    es_row = dict(es_row)
    exp_row = db_module._row(db_module.get_db(),
                             "SELECT wp_amount FROM expenses WHERE id=?", (es_row["expense_id"],))
    wp_amount = float(exp_row["wp_amount"] or 0) if exp_row else 0.0
    earned = round(wp_amount * float(es_row.get("split_pct") or 0) / 100, 2)
    paid = float(es_row.get("payroll_paid") or 0)
    status = db_module._payroll_status(earned, paid)
    badge = _payroll_badge_html(status, paid, es_row.get("payroll_paid_date") or "", sid=sid, kind="expense-splits")
    return f'<span id="exp-split-status-{sid}">{badge}</span>'


def _payroll_badge_html(status: str, paid: float, paid_date: str,
                        sid: int = 0, kind: str = "splits") -> str:
    """Return the inner HTML for a payroll paid-status cell."""
    if status in ("paid", "partial"):
        target      = f"split-status-{sid}" if kind == "splits" else f"exp-split-status-{sid}"
        toggle_url  = f"/payroll/{kind}/{sid}/toggle-paid"
        set_paid_url = f"/payroll/{kind}/{sid}/set-paid"
        date_val    = paid_date or ""
        badge_cls   = "status-released" if status == "paid" else "status-inv-pending"
        badge_txt   = (f"✓ Paid {paid_date}" if paid_date else "✓ Paid") if status == "paid" else f"Partial ${paid:,.2f}"
        return (
            f'<button type="button" class="status-badge {badge_cls} payroll-unmark-btn"'
            f' hx-post="{toggle_url}" hx-vals=\'{{"paid_date":""}}\''
            f' hx-target="#{target}" hx-swap="outerHTML"'
            f' title="Click to unmark paid">{badge_txt}</button>'
            f'<button type="button" class="btn-icon payroll-edit-btn"'
            f' onclick="payrollStartEdit(this)" title="Edit amount & date">✎</button>'
            f'<span class="payroll-editor" style="display:none">'
            f'<input type="number" step="0.01" min="0" value="{paid:.2f}"'
            f' name="amount" class="input-sm" style="width:160px">'
            f'<input type="date" name="paid_date" value="{date_val}" class="input-sm payroll-date-edit">'
            f'<button type="button" class="btn btn-primary btn-sm"'
            f' onclick="payrollSaveEdit(this,\'{set_paid_url}\',\'#{target}\')">Save</button>'
            f'<button type="button" class="btn btn-ghost btn-sm"'
            f' onclick="payrollCancelEdit(this)">Cancel</button>'
            f'</span>'
        )
    return '<span class="text-muted">—</span>'


# ---------------------------------------------------------------------------
# Payroll Runs
# ---------------------------------------------------------------------------

@app.route("/payroll/salary-employees")
@login_required
def salary_employees_list():
    employees = db_module.get_all_salary_employees()
    recipients = db_module.get_all_fee_recipients()
    return render_template("payroll_salary_employees.html",
                           employees=employees, recipients=recipients,
                           roles=ASSIGNEE_ROLES)


@app.route("/payroll/salary-employees", methods=["POST"])
@login_required
def salary_employee_create():
    name = request.form.get("name", "").strip()
    salary = db_module._f(request.form.get("monthly_salary"))
    company_name = request.form.get("company_name", "").strip() or None
    frid = request.form.get("fee_recipient_id") or None
    if frid:
        try:
            frid = int(frid)
        except ValueError:
            frid = None
    if name:
        db_module.create_salary_employee(name, salary, company_name, frid)
    employees = db_module.get_all_salary_employees()
    recipients = db_module.get_all_fee_recipients()
    return render_template("payroll_salary_employees.html",
                           employees=employees, recipients=recipients,
                           roles=ASSIGNEE_ROLES)


@app.route("/payroll/salary-employees/<int:emp_id>", methods=["PUT"])
@login_required
def salary_employee_update(emp_id: int):
    frid = request.form.get("fee_recipient_id") or None
    if frid:
        try:
            frid = int(frid)
        except ValueError:
            frid = None
    data = {
        "name": request.form.get("name", "").strip(),
        "company_name": request.form.get("company_name", "").strip() or None,
        "monthly_salary": db_module._f(request.form.get("monthly_salary")),
        "fee_recipient_id": frid,
        "active": request.form.get("active", "1") != "0",
    }
    db_module.update_salary_employee(emp_id, data)
    employees = db_module.get_all_salary_employees()
    recipients = db_module.get_all_fee_recipients()
    return render_template("payroll_salary_employees.html",
                           employees=employees, recipients=recipients,
                           roles=ASSIGNEE_ROLES)


@app.route("/payroll/salary-employees/<int:emp_id>", methods=["DELETE"])
@login_required
def salary_employee_delete(emp_id: int):
    db_module.delete_salary_employee(emp_id)
    employees = db_module.get_all_salary_employees()
    recipients = db_module.get_all_fee_recipients()
    return render_template("payroll_salary_employees.html",
                           employees=employees, recipients=recipients,
                           roles=ASSIGNEE_ROLES)


@app.route("/payroll/runs")
@login_required
def payroll_runs_list():
    runs = db_module.get_payroll_runs()
    return render_template("payroll_runs.html", runs=runs)


@app.route("/payroll/runs/new", methods=["GET"])
@login_required
def payroll_run_new_form():
    return render_template("payroll_runs.html",
                           runs=db_module.get_payroll_runs(),
                           show_new_form=True)


@app.route("/payroll/runs/new", methods=["POST"])
@login_required
def payroll_run_create():
    name = request.form.get("name", "").strip()
    if not name:
        name = "Untitled Run"
    period_from = _clean_date(request.form.get("period_from"))
    period_to = _clean_date(request.form.get("period_to"))
    pay_date = _clean_date(request.form.get("pay_date"))
    run_id = db_module.create_payroll_run(name, period_from, period_to, pay_date)
    return redirect(url_for("payroll_run_builder", run_id=run_id))


@app.route("/payroll/runs/<int:run_id>")
@login_required
def payroll_run_builder(run_id: int):
    data = db_module.get_run_builder_data(run_id)
    if not data:
        abort(404)
    return render_template("payroll_run_builder.html", **data)


@app.route("/payroll/runs/<int:run_id>/update", methods=["POST"])
@login_required
def payroll_run_update(run_id: int):
    run = db_module.get_payroll_run(run_id)
    if not run:
        abort(404)
    data = {
        "name": request.form.get("name", "").strip() or run["name"],
        "period_from": _clean_date(request.form.get("period_from")),
        "period_to": _clean_date(request.form.get("period_to")),
        "pay_date": _clean_date(request.form.get("pay_date")),
        "status": request.form.get("status", run.get("status", "draft")),
        "notes": request.form.get("notes", "").strip() or None,
    }
    db_module.update_payroll_run(run_id, data)
    run = db_module.get_payroll_run(run_id)
    return render_template("partials/_run_header.html", run=run)


@app.route("/payroll/runs/<int:run_id>", methods=["DELETE"])
@login_required
def payroll_run_delete(run_id: int):
    db_module.delete_payroll_run(run_id)
    runs = db_module.get_payroll_runs()
    return render_template("payroll_runs.html", runs=runs)


@app.route("/payroll/runs/<int:run_id>/mark-paid", methods=["POST"])
@login_required
def payroll_run_mark_paid(run_id: int):
    run = db_module.get_payroll_run(run_id)
    if not run:
        abort(404)
    pay_date = _clean_date(request.form.get("pay_date")) or run.get("pay_date") or ""
    db_module.mark_run_paid(run_id, pay_date)
    return redirect(url_for("payroll_run_builder", run_id=run_id))


@app.route("/payroll/runs/<int:run_id>/splits/<int:sid>/toggle", methods=["POST"])
@login_required
def run_split_toggle(run_id: int, sid: int):
    in_run = db_module.toggle_split_in_run(run_id, sid)
    data = db_module.get_run_builder_data(run_id)
    if not data:
        abort(404)
    # Find the payee that owns this split
    for payee in data["payees"]:
        for s in payee["tx_splits"]:
            if s["split_id"] == sid:
                return render_template("partials/_run_payee_totals.html",
                                       payee=payee, run=data["run"])
    return "", 204


@app.route("/payroll/runs/<int:run_id>/exp-splits/<int:sid>/toggle", methods=["POST"])
@login_required
def run_exp_split_toggle(run_id: int, sid: int):
    db_module.toggle_exp_split_in_run(run_id, sid)
    data = db_module.get_run_builder_data(run_id)
    if not data:
        abort(404)
    for payee in data["payees"]:
        for es in payee["exp_splits"]:
            if es["exp_split_id"] == sid:
                return render_template("partials/_run_payee_totals.html",
                                       payee=payee, run=data["run"])
    return "", 204


@app.route("/payroll/runs/<int:run_id>/salaries/<int:eid>/set", methods=["POST"])
@login_required
def run_salary_set(run_id: int, eid: int):
    try:
        amount = float(request.form.get("amount", 0))
    except (TypeError, ValueError):
        amount = 0.0
    db_module.set_run_salary_override(run_id, eid, amount)
    # Return updated salary row fragment
    data = db_module.get_run_builder_data(run_id)
    if not data:
        abort(404)
    for se in data["salary_employees"]:
        if se["id"] == eid:
            return f'<span id="sal-amount-{eid}">${amount:,.2f}</span>'
    return f'<span id="sal-amount-{eid}">${amount:,.2f}</span>'


@app.route("/payroll/runs/<int:run_id>/employees/<path:payee_name>/print")
@login_required
def run_employee_print(run_id: int, payee_name: str):
    data = db_module.get_run_builder_data(run_id)
    if not data:
        abort(404)
    # Find the payee
    payee = None
    for p in data["payees"]:
        if p["name"] == payee_name:
            payee = p
            break
    # Also check salary-only employees
    salary_emp = None
    if payee is None:
        for se in data["salary_employees"]:
            if se["name"] == payee_name and se.get("fee_recipient_id") is None:
                salary_emp = se
                break
    if payee is None and salary_emp is None:
        abort(404)
    return render_template("payroll_employee_print.html",
                           run=data["run"],
                           payee=payee,
                           salary_emp=salary_emp)


@app.route("/payroll/runs/<int:run_id>/export.xlsx")
@login_required
def payroll_run_export(run_id: int):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    data = db_module.get_run_builder_data(run_id)
    if not data:
        abort(404)
    run = data["run"]

    wb = openpyxl.Workbook()

    # ---- Summary tab ----
    ws = wb.active
    ws.title = "Payroll Summary"
    header_font = Font(bold=True)
    money_fmt = '#,##0.00'

    ws.append(["Name", "Salary", "Fee Earnings", "Expenses", "Total"])
    for cell in ws[1]:
        cell.font = header_font

    grand_salary = grand_fees = grand_exp = grand_total = 0.0

    # Salary-only employees first
    for se in data["salary_employees"]:
        if se.get("fee_recipient_id") is not None:
            continue  # will be shown in payee section
        amt = float(se.get("run_amount") or se.get("monthly_salary") or 0)
        ws.append([se["name"], amt, 0, 0, amt])
        for cell in ws[ws.max_row][1:]:
            cell.number_format = money_fmt
        grand_salary += amt
        grand_total += amt

    # Fee recipients
    for payee in data["payees"]:
        sel_tx = payee["selected_tx_total"]
        sel_exp = payee["selected_exp_total"]
        sal = payee["salary_amount"]
        total = round(sel_tx + sel_exp + sal, 2)
        ws.append([payee["name"], sal, sel_tx, sel_exp, total])
        for cell in ws[ws.max_row][1:]:
            cell.number_format = money_fmt
        grand_salary += sal
        grand_fees += sel_tx
        grand_exp += sel_exp
        grand_total += total

    # Totals row
    ws.append(["TOTAL", grand_salary, grand_fees, grand_exp, grand_total])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        if cell.column > 1:
            cell.number_format = money_fmt

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    # ---- Per-payee tabs ----
    for payee in data["payees"]:
        sel_tx = [s for s in payee["tx_splits"] if s["in_run"]]
        sel_exp = [s for s in payee["exp_splits"] if s["in_run"]]
        if not sel_tx and not sel_exp and not payee["salary_amount"]:
            continue

        tab_name = payee["name"][:31]
        ws2 = wb.create_sheet(title=tab_name)

        # Header block
        ws2.append(["Name:", payee["name"]])
        ws2.append(["Company:", payee.get("company_name") or ""])
        ws2.append(["Pay Period:",
                    f"{run.get('period_from','') or ''} – {run.get('period_to','') or ''}"])
        ws2.append(["Pay Date:", run.get("pay_date") or ""])
        ws2.append([])

        if sel_tx:
            ws2.append(["Job #", "Insured", "Check #", "Date", "Type", "Split %", "Earned"])
            for cell in ws2[ws2.max_row]:
                cell.font = header_font
            for s in sel_tx:
                ws2.append([
                    s["job_number"], s["insured_name"], s["check_number"],
                    s["date"], s["type"],
                    s["split_pct"], s["earned"],
                ])
                ws2.cell(ws2.max_row, 7).number_format = money_fmt
            ws2.append([])

        if sel_exp:
            ws2.append(["Job #", "Insured", "Date", "Vendor", "Split %", "Owed"])
            for cell in ws2[ws2.max_row]:
                cell.font = header_font
            for es in sel_exp:
                ws2.append([
                    es["job_number"], es["insured_name"], es["date"],
                    es["vendor"], es["split_pct"], es["owed"],
                ])
                ws2.cell(ws2.max_row, 6).number_format = money_fmt
            ws2.append([])

        # Summary footer
        fee_sub = payee["selected_tx_total"]
        exp_sub = payee["selected_exp_total"]
        sal = payee["salary_amount"]
        ws2.append(["Fee Subtotal:", fee_sub])
        ws2.cell(ws2.max_row, 2).number_format = money_fmt
        ws2.append(["Expenses Subtotal:", exp_sub])
        ws2.cell(ws2.max_row, 2).number_format = money_fmt
        ws2.append(["Base Salary:", sal])
        ws2.cell(ws2.max_row, 2).number_format = money_fmt
        ws2.append(["GRAND TOTAL:", round(fee_sub + exp_sub + sal, 2)])
        ws2.cell(ws2.max_row, 1).font = header_font
        ws2.cell(ws2.max_row, 2).font = header_font
        ws2.cell(ws2.max_row, 2).number_format = money_fmt

        for col in ws2.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)

    # ---- Salary-only tabs ----
    for se in data["salary_employees"]:
        if se.get("fee_recipient_id") is not None:
            continue
        amt = float(se.get("run_amount") or se.get("monthly_salary") or 0)
        tab_name = (se["name"] + " (Salary)")[:31]
        ws3 = wb.create_sheet(title=tab_name)
        ws3.append(["Name:", se["name"]])
        ws3.append(["Company:", se.get("company_name") or ""])
        ws3.append(["Pay Period:",
                    f"{run.get('period_from','') or ''} – {run.get('period_to','') or ''}"])
        ws3.append(["Pay Date:", run.get("pay_date") or ""])
        ws3.append([])
        ws3.append(["Salary:", amt])
        ws3.cell(ws3.max_row, 2).number_format = money_fmt
        ws3.cell(ws3.max_row, 1).font = Font(bold=True)
        ws3.cell(ws3.max_row, 2).font = Font(bold=True)
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = run["name"].replace(" ", "_").replace("/", "-")
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )


# ---------------------------------------------------------------------------
# CAT File Import
# ---------------------------------------------------------------------------

TMP_DIR = os.path.join(os.path.dirname(__file__), "tmp")


def _save_parsed(data: dict) -> str:
    """Write parsed data to a temp JSON file. Returns the key (UUID)."""
    os.makedirs(TMP_DIR, exist_ok=True)
    key = str(uuid.uuid4())
    path = os.path.join(TMP_DIR, f"{key}.json")
    with open(path, "w") as f:
        json.dump(data, f, default=str)
    return key


def _load_parsed(key: str) -> dict | None:
    if not key:
        return None
    path = os.path.join(TMP_DIR, f"{key}.json")
    if not os.path.exists(path):
        return None
    with open(path) as f:
        return json.load(f)


def _delete_parsed(key: str) -> None:
    if not key:
        return
    path = os.path.join(TMP_DIR, f"{key}.json")
    try:
        os.remove(path)
    except OSError:
        pass


@app.route("/import", methods=["GET"])
@login_required
def import_page():
    return render_template("import.html", parsed=None, parse_errors=[])


@app.route("/import", methods=["POST"])
@login_required
def import_upload():
    f = request.files.get("xlsx_file")
    if not f or not f.filename:
        flash("Please select an Excel file to upload.", "error")
        return redirect(url_for("import_page"))

    if not f.filename.lower().endswith(".xlsx"):
        flash("Only .xlsx files are supported.", "error")
        return redirect(url_for("import_page"))

    try:
        import openpyxl
        import import_cat as ic
        wb = openpyxl.load_workbook(f, data_only=True)
        parsed = ic.parse_workbook(wb)
    except Exception as e:
        flash(f"Failed to parse file: {e}", "error")
        return redirect(url_for("import_page"))

    key = _save_parsed(parsed)
    session["import_key"] = key
    return redirect(url_for("import_review"))


@app.route("/import/review", methods=["GET"])
@login_required
def import_review():
    key = session.get("import_key")
    parsed = _load_parsed(key) if key else None
    if not parsed:
        flash("No import data found. Please upload a file.", "error")
        return redirect(url_for("import_page"))
    vendors = db_module.get_all_vendors()
    return render_template(
        "import.html",
        parsed=parsed,
        parse_errors=parsed.get("parse_errors", []),
        tx_types=TRANSACTION_TYPES,
        assignee_roles=ASSIGNEE_ROLES,
        vendors=vendors,
    )


@app.route("/import/commit", methods=["POST"])
@login_required
def import_commit():
    key = session.get("import_key")
    parsed = _load_parsed(key) if key else None
    if not parsed:
        flash("Import session expired. Please upload again.", "error")
        return redirect(url_for("import_page"))

    try:
        claim_id = _do_import(request.form, parsed)
    except Exception as e:
        flash(f"Import failed: {e}", "error")
        return redirect(url_for("import_review"))

    _delete_parsed(key)
    session.pop("import_key", None)
    flash("Import complete! Review and fill in fee details below.", "success")
    return redirect(url_for("claim_detail", claim_id=claim_id))


def _do_import(form, parsed: dict) -> int:
    """Read flat form fields and write to DB. Returns new claim_id."""
    f = form

    # --- Claim ---
    claim_data = {
        "insured_name":  f.get("claim_insured_name", "").strip(),
        "carrier":       f.get("claim_carrier", "").strip(),
        "contract_pct":  db_module._f(f.get("claim_contract_pct")),
        "job_number":    f.get("claim_job_number", "").strip(),
        "claim_number":  f.get("claim_claim_number", "").strip(),
        "mortgage_company": None,
        "notes": None,
        "proc_method_carrier": None,
        "proc_method_draw":    None,
        "proc_method_escrow":  None,
    }
    claim_id = db_module.create_claim(claim_data)

    # --- Limits ---
    limit_count = int(f.get("limit_count", 0))
    for i in range(limit_count):
        if not f.get(f"limit_{i}_include"):
            continue
        data = {
            "coverage_type":  f.get(f"limit_{i}_coverage_type", "").strip(),
            "base_limit":     db_module._f(f.get(f"limit_{i}_base_limit")),
            "extended_limit": db_module._f(f.get(f"limit_{i}_extended_limit")),
        }
        db_module.ensure_coverage_type(data["coverage_type"])
        db_module.create_limit(claim_id, data)

    # --- Assignees ---
    assignee_count = int(f.get("assignee_count", 0))
    for i in range(assignee_count):
        if not f.get(f"assignee_{i}_include"):
            continue
        data = {
            "role":      f.get(f"assignee_{i}_role", "other"),
            "name":      f.get(f"assignee_{i}_name", "").strip(),
            "split_pct": db_module._f(f.get(f"assignee_{i}_split_pct")),
            "sort_order": i,
            "recipient_id": None,
        }
        db_module.create_assignee(claim_id, data)

    # --- Transactions ---
    tx_count = int(f.get("tx_count", 0))
    for i in range(tx_count):
        if not f.get(f"tx_{i}_include"):
            continue
        tx_type = f.get(f"tx_{i}_type", "Carrier")
        tx_data = {
            "sequence_number": f.get(f"tx_{i}_seq") or None,
            "check_number":    f.get(f"tx_{i}_check_number", "").strip() or None,
            "date":            f.get(f"tx_{i}_check_date", "").strip(),
            "received_date":   f.get(f"tx_{i}_received_date", "").strip(),
            "amount":          db_module._f(f.get(f"tx_{i}_amount")),
            "type":            tx_type,
            "payer":           f.get(f"tx_{i}_payer", "").strip(),
            "payees_text":     f.get(f"tx_{i}_payees_text", "").strip(),
            "invoice_number":  f.get(f"tx_{i}_invoice_number", "").strip(),
            # Fee fields intentionally 0 at import — user fills in post-import
            "fee_owed": 0.0, "balance": 0.0, "deferred": 0.0, "recouped": 0.0,
            "fee_collected": 0.0, "reimbursed": 0.0, "total_collected": 0.0,
            "unpaid_payee_expense": 0.0, "outstanding_expense": 0.0,
            "ott": None, "notes": "", "check_id": None,
            "endorsed": False, "void": bool(f.get(f"tx_{i}_void")),
            "linked_escrow_id": None, "inv_for": "insured", "inv_vendor_id": None,
            "proc_method": None, "proc_fee_invoice_id": None,
        }
        tx_id = db_module.create_transaction(claim_id, tx_data)

        # If the Excel sheet had its own SPLITS section, replace the auto-cloned
        # assignee splits with the per-check splits from the file.
        parsed_txs = parsed.get("transactions") or []
        parsed_tx_splits = parsed_txs[i]["splits"] if i < len(parsed_txs) else []
        if parsed_tx_splits:
            db_module.get_db().execute(
                "DELETE FROM transaction_splits WHERE transaction_id=?", (tx_id,))
            db_module.get_db().commit()
            for sp in parsed_tx_splits:
                db_module.create_split(tx_id, sp)

        # For fee invoices, create the payment record if checked and amount > 0
        if tx_type == "Fee Invoice" and f.get(f"tx_{i}_pay_include"):
            pay_amt = db_module._f(f.get(f"tx_{i}_pay_amount"))
            if pay_amt and pay_amt > 0:
                db_module.create_invoice_payment(tx_id, {
                    "date":      f.get(f"tx_{i}_received_date", "").strip(),
                    "amount":    pay_amt,
                    "method":    "qb",
                    "reference": f.get(f"tx_{i}_pay_reference", "").strip(),
                })

        # Coverage allocations
        cov_count = int(f.get(f"tx_{i}_cov_count", 0))
        for j in range(cov_count):
            if not f.get(f"tx_{i}_cov_{j}_include"):
                continue
            cov_type = f.get(f"tx_{i}_cov_{j}_coverage_type", "").strip()
            if cov_type:
                db_module.ensure_coverage_type(cov_type)
            db_module.create_coverage(tx_id, {
                "coverage_type": cov_type,
                "amount": db_module._f(f.get(f"tx_{i}_cov_{j}_amount")),
            })

        # Disbursements
        disb_count = int(f.get(f"tx_{i}_disb_count", 0))
        for j in range(disb_count):
            if not f.get(f"tx_{i}_disb_{j}_include"):
                continue
            vendor_id = f.get(f"tx_{i}_disb_{j}_vendor_id") or None
            recipient_name = f.get(f"tx_{i}_disb_{j}_recipient_name", "").strip()
            if not vendor_id and f.get(f"tx_{i}_disb_{j}_recipient_type") == "vendor":
                vendor_id = db_module.ensure_vendor_by_name(recipient_name)
            fee_pct_raw = db_module._fopt(f.get(f"tx_{i}_disb_{j}_fee_pct"))
            fee_owed_val  = db_module._f(f.get(f"tx_{i}_disb_{j}_fee_owed"))
            fee_defer_val = db_module._f(f.get(f"tx_{i}_disb_{j}_fee_deferred"))
            # Mirror the claim's "Fee Rcvd" toggle: set fee_collected = fee_owed - fee_deferred
            fee_received = bool(f.get(f"tx_{i}_disb_{j}_fee_received"))
            fee_recoup_val = db_module._f(f.get(f"tx_{i}_disb_{j}_fee_recouped"))
            if fee_received and fee_pct_raw and fee_owed_val > 0:
                fee_collected_val = round(max(0.0, fee_owed_val - fee_defer_val + fee_recoup_val), 2)
            else:
                fee_collected_val = 0.0
            db_module.create_disbursement(tx_id, {
                "sort_order":     j,
                "recipient_type": f.get(f"tx_{i}_disb_{j}_recipient_type", "insured"),
                "vendor_id":      vendor_id,
                "recipient_name": recipient_name,
                "amount":         db_module._f(f.get(f"tx_{i}_disb_{j}_amount")),
                "fee_applies":    bool(fee_pct_raw),
                "fee_pct":        fee_pct_raw,
                "fee_owed":       fee_owed_val,
                "fee_collected":  fee_collected_val,
                "fee_deferred":   fee_defer_val,
                "fee_recouped":   db_module._f(f.get(f"tx_{i}_disb_{j}_fee_recouped")),
                "client_received": 1 if f.get(f"tx_{i}_disb_{j}_client_received") else 0,
                "use_check_splits": True, "notes": "",
            })

    # --- Expenses ---
    exp_count = int(f.get("exp_count", 0))
    for i in range(exp_count):
        if not f.get(f"exp_{i}_include"):
            continue
        exp_payee = f.get(f"exp_{i}_payee_name", "").strip()
        exp_id = db_module.create_expense(claim_id, {
            "invoice_date":      f.get(f"exp_{i}_invoice_date", "").strip(),
            "payee_name":        exp_payee,
            "vendor_id":         db_module.ensure_vendor_by_name(exp_payee),
            "invoice_amount":    db_module._f(f.get(f"exp_{i}_invoice_amount")),
            "responsible_party": f.get(f"exp_{i}_responsible_party", "").strip(),
            "unpaid_to_payee":   0.0,
            "client_outstanding": 0.0,
            "wp_outstanding":    0.0,
            "reason":            f.get(f"exp_{i}_reason", "").strip(),
            "invoice_number":    f.get(f"exp_{i}_invoice_number", "").strip(),
            "client_amount":     db_module._f(f.get(f"exp_{i}_client_amount")),
            "wp_amount":         db_module._f(f.get(f"exp_{i}_wp_amount")),
        })

        # If the expense sheet had its own SPLITS section, replace auto-cloned splits.
        parsed_exps = parsed.get("expenses") or []
        parsed_exp_splits = parsed_exps[i]["splits"] if i < len(parsed_exps) else []
        if parsed_exp_splits:
            db_module.get_db().execute(
                "DELETE FROM expense_splits WHERE expense_id=?", (exp_id,))
            db_module.get_db().commit()
            for sp in parsed_exp_splits:
                db_module.create_expense_split(exp_id, sp)

        # Create expense_payment records for amounts already paid to payee
        client_paid = db_module._f(f.get(f"exp_{i}_client_paid"))
        if client_paid and client_paid > 0:
            db_module.create_expense_payment(exp_id, {
                "paid_by": "client",
                "amount":  client_paid,
                "date":    f.get(f"exp_{i}_invoice_date", "").strip(),
            })
        wp_paid = db_module._f(f.get(f"exp_{i}_wp_paid"))
        if wp_paid and wp_paid > 0:
            db_module.create_expense_payment(exp_id, {
                "paid_by": "wp",
                "amount":  wp_paid,
                "date":    f.get(f"exp_{i}_invoice_date", "").strip(),
            })

    return claim_id


# ---------------------------------------------------------------------------
# Global Status
# ---------------------------------------------------------------------------

@app.route("/status")
@login_required
def global_status():
    data = db_module.get_global_status()
    return render_template("status.html", **data)


@app.route("/status/export.csv")
@login_required
def status_export():
    data = db_module.get_global_status()
    adjuster = request.args.get("adjuster", "").strip()

    def _adj_match(claim_id):
        if not adjuster:
            return True
        return adjuster in data["claim_adjusters"].get(claim_id, [])

    pending_checks       = [r for r in data["pending_checks"]       if _adj_match(r["claim_id"])]
    open_escrow          = [r for r in data["open_escrow"]          if _adj_match(r["claim_id"])]
    unrecouped           = [r for r in data["unrecouped_deferred"]  if _adj_match(r["claim_id"])]
    pending_expenses     = [r for r in data["pending_expenses"]     if _adj_match(r["claim_id"])]
    pending_fee_invoices = [r for r in data["pending_fee_invoices"] if _adj_match(r["claim_id"])]

    out = io.StringIO()
    writer = csv.writer(out)

    writer.writerow(["PENDING CHECKS"])
    writer.writerow(["Insured", "Job #", "Check #", "Type", "Date", "Amount",
                     "Fee Rcvd", "Client Rcvd", "Adjusters", "Notes"])
    for r in pending_checks:
        fee_ok = not r.get("fee_not_received_count") or r["fee_not_received_count"] == 0
        client_ok = not r.get("pending_client_count") or r["pending_client_count"] == 0
        adjs = " | ".join(data["claim_adjusters"].get(r["claim_id"], []))
        writer.writerow([
            r["insured_name"], r["job_number"] or "", r["check_number"] or "",
            r["type"], r["check_date"] or "", r["amount"],
            "Yes" if (not r.get("fee_applicable_count") or fee_ok) else "Pending",
            "Yes" if client_ok else f"{r['pending_client_count']} pending",
            adjs, r.get("status_notes") or "",
        ])

    writer.writerow([])
    writer.writerow(["MONEY IN ESCROW"])
    writer.writerow(["Insured", "Job #", "Claim #", "Escrowed", "Drawn", "Net in Escrow",
                     "Adjusters", "Notes"])
    for r in open_escrow:
        adjs = " | ".join(data["claim_adjusters"].get(r["claim_id"], []))
        writer.writerow([
            r["insured_name"], r["job_number"] or "", r["claim_number"] or "",
            r["total_escrowed"], r["total_drawn"], r["net_in_escrow"],
            adjs, r.get("status_notes") or "",
        ])

    writer.writerow([])
    writer.writerow(["UNRECOUPED DEFERRED FEES"])
    writer.writerow(["Insured", "Job #", "Claim #", "Deferred", "Recouped", "Outstanding",
                     "Adjusters"])
    for r in unrecouped:
        adjs = " | ".join(data["claim_adjusters"].get(r["claim_id"], []))
        writer.writerow([
            r["insured_name"], r["job_number"] or "", r["claim_number"] or "",
            r["total_deferred"], r["total_recouped"], r["unrecouped"], adjs,
        ])

    writer.writerow([])
    writer.writerow(["PENDING FEE INVOICES"])
    writer.writerow(["Insured", "Job #", "Invoice #", "Date", "Invoice Amt",
                     "Collected", "Balance Due", "Adjusters"])
    for r in pending_fee_invoices:
        adjs = " | ".join(data["claim_adjusters"].get(r["claim_id"], []))
        writer.writerow([
            r["insured_name"], r["job_number"] or "", r["invoice_number"] or "",
            r["date"] or "", r["amount"], r["total_collected"], r["balance_due"], adjs,
        ])

    writer.writerow([])
    writer.writerow(["PENDING EXPENSES"])
    writer.writerow(["Insured", "Job #", "Date", "Payee", "Reason", "Responsible Party",
                     "Invoice Amount", "Owed to Payee", "Reimburse Owed", "WP Reimb Owed",
                     "Adjusters", "Notes"])
    for r in pending_expenses:
        adjs = " | ".join(data["claim_adjusters"].get(r["claim_id"], []))
        writer.writerow([
            r["insured_name"], r["job_number"] or "", r["invoice_date"] or "",
            r["payee_name"] or "", r["reason"] or "", r["responsible_party"] or "",
            r["invoice_amount"], r["outstanding"], r["reimburse_owed"], r["wp_reimburse_owed"],
            adjs, r.get("status_notes") or "",
        ])

    out.seek(0)
    suffix = f"_{adjuster.replace(' ', '_')}" if adjuster else ""
    filename = f"status{suffix}_{date.today().isoformat()}.csv"
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/status/tx/<int:tx_id>/note", methods=["POST"])
@login_required
def status_tx_note(tx_id: int):
    notes = request.form.get("notes", "").strip()
    db_module.set_status_note("tx", tx_id, notes)
    return render_template("partials/_status_note.html",
                           notes=notes, save_url=url_for("status_tx_note", tx_id=tx_id))


@app.route("/status/claim/<int:claim_id>/note", methods=["POST"])
@login_required
def status_claim_note(claim_id: int):
    notes = request.form.get("notes", "").strip()
    db_module.set_status_note("claim", claim_id, notes)
    return render_template("partials/_status_note.html",
                           notes=notes, save_url=url_for("status_claim_note", claim_id=claim_id))


@app.route("/status/expense/<int:exp_id>/note", methods=["POST"])
@login_required
def status_expense_note(exp_id: int):
    notes = request.form.get("notes", "").strip()
    db_module.set_status_note("expense", exp_id, notes)
    return render_template("partials/_status_note.html",
                           notes=notes, save_url=url_for("status_expense_note", exp_id=exp_id))


@app.route("/status/tx/<int:tx_id>/follow-up", methods=["POST"])
@login_required
def status_tx_followup(tx_id: int):
    date_str = request.form.get("follow_up_date", "").strip()
    db_module.set_follow_up_date("tx", tx_id, date_str)
    return render_template("partials/_status_followup.html",
                           follow_up_date=date_str,
                           save_url=url_for("status_tx_followup", tx_id=tx_id))


@app.route("/status/claim/<int:claim_id>/follow-up", methods=["POST"])
@login_required
def status_claim_followup(claim_id: int):
    date_str = request.form.get("follow_up_date", "").strip()
    db_module.set_follow_up_date("claim", claim_id, date_str)
    return render_template("partials/_status_followup.html",
                           follow_up_date=date_str,
                           save_url=url_for("status_claim_followup", claim_id=claim_id))


@app.route("/status/expense/<int:exp_id>/follow-up", methods=["POST"])
@login_required
def status_expense_followup(exp_id: int):
    date_str = request.form.get("follow_up_date", "").strip()
    db_module.set_follow_up_date("expense", exp_id, date_str)
    return render_template("partials/_status_followup.html",
                           follow_up_date=date_str,
                           save_url=url_for("status_expense_followup", exp_id=exp_id))


# ---------------------------------------------------------------------------
# App startup
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    db_module.init_db()
    print(f"Database: {config.DB_PATH}")
    print("Run seed.py first if this is a fresh install.")
    app.run(debug=config.DEBUG, host="0.0.0.0", port=5001)
